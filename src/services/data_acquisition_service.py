# -*- coding: utf-8 -*-
"""Model-planned, auditable multi-source financial-data acquisition packages."""

from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta, timezone
import csv
import hashlib
import json
import logging
import os
from pathlib import Path
import re
import shutil
import subprocess
import tempfile
from typing import Any, Callable, Dict, List, Optional
from urllib.parse import unquote, urlparse
import zipfile

from json_repair import repair_json
from openpyxl import Workbook
import requests

from src.config import get_config
from src.services.essay_analysis_service import DEFAULT_ESSAY_MODEL, EssayDailyReportService
from src.services.cninfo_announcement_service import CninfoAnnouncementService
from src.services.financial_data_service import FinancialDataService, FinancialDataValidationError
from src.services.investment_monitor_service import InvestmentMonitorService

logger = logging.getLogger(__name__)

MAX_TASKS = 12
MAX_ROWS_PER_DATASET = 2000
MAX_RESEARCH_REPORT_RECALL_ROWS = max(
    5000, min(int(os.getenv("DATA_ACQUISITION_REPORT_RECALL_MAX_ROWS", "100000")), 200000)
)
MAX_RESEARCH_REPORT_RESULTS = max(
    20, min(int(os.getenv("DATA_ACQUISITION_REPORT_MAX_RESULTS", "500")), MAX_ROWS_PER_DATASET)
)
_RESOURCE_RE = re.compile(r"^[a-z][a-z0-9_]{0,63}$")

# The planner sees the most useful productized endpoints. The executor still supports
# every syntactically valid Tushare api_name so the user's full Pro entitlement remains callable.
TUSHARE_CATALOG = {
    "daily": "A股日线行情", "daily_basic": "每日估值与换手", "adj_factor": "复权因子",
    "stk_factor": "技术因子", "moneyflow": "个股资金流", "moneyflow_hsgt": "沪深港通资金",
    "margin_detail": "融资融券明细", "hk_hold": "沪深港通持股", "cyq_perf": "筹码胜率与成本",
    "cyq_chips": "筹码分布", "income": "利润表", "balancesheet": "资产负债表",
    "cashflow": "现金流量表", "fina_indicator": "财务指标", "forecast": "业绩预告",
    "express": "业绩快报", "dividend": "分红送股", "fina_audit": "财务审计意见",
    "pledge_stat": "股权质押", "share_float": "限售股解禁", "stk_holdertrade": "股东增减持",
    "repurchase": "股份回购", "top10_holders": "十大股东", "top10_floatholders": "十大流通股东",
    "block_trade": "大宗交易", "top_list": "龙虎榜", "top_inst": "龙虎榜机构明细",
    "concept": "概念分类", "concept_detail": "概念成分", "index_daily": "指数日线",
    "index_weight": "指数成分权重", "ths_hot": "同花顺热榜", "dc_hot": "东方财富热榜",
    "research_report": "可下载券商研报（个股/行业/策略）", "report_rc": "卖方盈利预测与研报索引",
    "broker_recommend": "券商月度金股", "major_news": "长篇财经新闻",
    "news": "实时快讯", "cctv_news": "新闻联播", "anns_d": "公告数据",
    "trade_cal": "交易日历", "stock_basic": "上市公司基础资料", "namechange": "证券曾用名",
    "new_share": "新股上市", "bak_basic": "备用基础行情", "stk_limit": "涨跌停价格",
    "limit_list_d": "涨跌停榜单", "suspend_d": "停复牌", "disclosure_date": "财报披露计划",
    "moneyflow_cnt_ths": "同花顺概念资金流", "moneyflow_ind_ths": "同花顺行业资金流",
    "limit_list_ths": "同花顺涨停/炸板/跌停池", "margin": "全市场融资融券汇总",
    "stk_nineturn": "神奇九转", "stk_managers": "上市公司管理层名录",
}

TYC_COMMANDS = {
    "company_full": [],
    "registration": ["company", "registration-info"],
    "company_search": ["company", "companies"],
    "company_profile": ["company", "profile"],
    "company_financials": ["company", "financial-data"],
    "shareholders": ["company", "shareholder-info"],
    "actual_controller": ["company", "actual-controller"],
    "risk_overview": ["risk", "overview"],
    "operation_credit": ["operation", "credit-evaluation"],
    "ipr_score": ["intellectual_property", "ipr-score"],
    "historical_overview": ["history", "historical-overview"],
}

TYC_FULL_COMMANDS = {
    "registration": ["company", "registration-info"],
    "profile": ["company", "profile"],
    "listing": ["company", "listing-info"],
    "financials": ["company", "financial-data"],
    "annual_reports": ["company", "annual-reports"],
    "shareholders": ["company", "shareholder-info"],
    "actual_controller": ["company", "actual-controller"],
    "beneficial_owners": ["company", "beneficial-owners"],
    "external_investments": ["company", "external-investments"],
    "key_personnel": ["company", "key-personnel"],
    "relation_graph": ["company", "relation-graph"],
    "risk_overview": ["risk", "overview"],
    "judicial_cases": ["risk", "judicial-case"],
    "administrative_penalties": ["risk", "administrative-penalty"],
    "business_exceptions": ["risk", "business-exception"],
    "credit": ["operation", "credit-evaluation"],
    "news_sentiment": ["operation", "news-sentiment"],
    "bidding": ["operation", "bidding-info"],
    "suppliers_customers": ["operation", "suppliers-and-customers"],
    "products": ["operation", "products-info"],
    "ipr_score": ["intellectual_property", "ipr-score"],
    "patents": ["intellectual_property", "patent-info"],
    "trademarks": ["intellectual_property", "trademark-info"],
    "historical_overview": ["history", "historical-overview"],
}

_FILE_REQUEST_RE = re.compile(
    r"(?:下载|打包|附上|包含|提供).{0,10}(?:文件|附件|PDF|原文|研报|报告)"
    r"|(?:文件|附件|PDF|原文|研报|报告).{0,10}(?:下载|打包|一并)",
    re.I,
)
MAX_DOWNLOAD_FILES = 100
MAX_FILE_BYTES = 50 * 1024 * 1024
MAX_PACKAGE_FILE_BYTES = 500 * 1024 * 1024

PLANNER_SYSTEM = """你是财经数据产品的数据调度规划器。把用户需求转成严格 JSON，不回答投资建议。
只能使用 capability_catalog 中的数据源和资源。只获取用户明确要求的数据维度，不得因为目录中存在其他能力就一并添加；用户要求“只/仅/不要”时必须严格排除未要求渠道。多维综合请求再覆盖行情、财务、资金、公告、新闻、机构观点、知识星球和工商风险中真正相关的维度，避免重复。
股票代码必须使用 Tushare 格式，如 603306.SH、300476.SZ。日期参数使用 YYYYMMDD；本地库日期使用 YYYY-MM-DD。
公告必须使用 cninfo.announcements 现场调用巨潮接口重新检索，不得使用本地公告缓存。Tushare 数据必须按维度直接调用对应 api_name；研报 PDF 库使用 research_report，盈利预测使用 report_rc，新闻使用 news/major_news，不能用 monitor.events 代替。不要向新闻接口传 ts_code。
研报主题请求必须把用户原话中的主题拆成 scope.keywords 和 research_report.params.topics；“或/或者”使用 keyword_mode=any。时间范围必须据 today 精确计算，不得用接口默认的20天代替“最近两年”。“最好是深度研究”设 depth_preference=prefer，“只要/必须是深度研究”设为strict，并设 ai_filter=true。
例：“下载最近两年低空经济或无人机方向的深度研报” => scope.keywords=["低空经济","无人机"]，scope.start_date=当天向前两年，scope.end_date=today，task.params={"topics":["低空经济","无人机"],"keyword_mode":"any","depth_preference":"prefer","ai_filter":true,...}，include_files=true。
知识星球使用 created_from、created_to、query 或 symbol；天眼查的 company_name 可以是单个名称或名称数组。
必须输出全局 scope：symbols、company_names、keywords、start_date、end_date、market_wide。除非用户明确要求全市场，market_wide 必须为 false。
每个任务必须包含 source、resource、label、reason、params、fields。params 只能包含该渠道实际查询参数，不能用空参数拉全量。最多12个任务。
天眼查要求“全面/全景/尽调”时使用 tianyancha.company_full；它会现场覆盖工商、股权、实控人、财务、司法风险、经营、知识产权和历史信息。
若用户要求下载、文件、附件、PDF或原文文件，include_files 必须为 true；否则为 false。
输出结构：{"title":"...","objective":"...","scope":{"symbols":[],"company_names":[],"keywords":[],"start_date":"","end_date":"","market_wide":false},"tasks":[...],"include_files":false,"output_formats":["json","csv","xlsx","zip"],"caveats":[]}。
不得输出 Markdown 或 JSON 外文字。"""


class DataAcquisitionError(RuntimeError):
    """Safe error surfaced by the acquisition workbench."""


def _json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, dict):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_json_safe(item) for item in value]
    if hasattr(value, "item"):
        try:
            return _json_safe(value.item())
        except (TypeError, ValueError):
            pass
    return str(value)


_REPORT_TOPIC_ALIASES: Dict[str, List[str]] = {
    "低空经济": ["低空经济", "低空", "通用航空", "通航", "eVTOL", "飞行汽车", "城市空中交通", "UAM"],
    "无人机": ["无人机", "无人飞行器", "UAV", "无人航空器", "无人航空", "飞控"],
}
_REPORT_INTERNAL_PARAMS = {
    "topics", "keywords", "keyword", "query", "keyword_mode", "depth_preference",
    "ai_filter", "max_results", "strict_depth",
}
_REPORT_TOPIC_STOPWORDS = {
    "研报", "报告", "研究", "深度", "深度研究", "主题", "方向", "最近", "近期",
    "下载", "打包", "最好", "都是", "我要", "筛选", "数据",
}


def _split_search_terms(value: Any) -> List[str]:
    """Normalize planner/user topic values without treating an OR phrase as one literal."""
    values = value if isinstance(value, (list, tuple, set)) else [value]
    terms: List[str] = []
    for item in values:
        text = str(item or "").strip()
        if not text:
            continue
        for part in re.split(r"(?:或者|以及|和|\bor\b|[\s,，、;；|/]+)", text, flags=re.I):
            cleaned = re.sub(r"^(?:关于|聚焦|围绕|主题(?:要是|是|为)?)+", "", part.strip())
            cleaned = re.sub(r"(?:方向|主题|板块|产业|行业)$", "", cleaned).strip()
            if not cleaned or cleaned in _REPORT_TOPIC_STOPWORDS or len(cleaned) > 40:
                continue
            terms.append(cleaned)
    return list(dict.fromkeys(terms))


def _extract_report_topics(request_text: str) -> List[str]:
    """Extract a conservative topic set as a guard when the LLM drops constraints."""
    text = str(request_text or "").strip()
    topics: List[str] = []
    for canonical, aliases in _REPORT_TOPIC_ALIASES.items():
        if any(alias.lower() in text.lower() for alias in aliases):
            topics.append(canonical)
    patterns = (
        r"主题(?:要是|是|为)?(.{1,80}?)(?:[，。；;]|最好|优先|最近|近\s*\d|过去|$)",
        r"(?:关于|聚焦|围绕)(.{1,60}?)(?:的)?(?:研报|报告|[，。；;]|$)",
    )
    for pattern in patterns:
        for match in re.finditer(pattern, text, re.I):
            topics.extend(_split_search_terms(match.group(1)))
    return list(dict.fromkeys(topics))


def _chinese_number(value: str) -> Optional[int]:
    raw = str(value or "").strip()
    if raw.isdigit():
        return int(raw)
    mapping = {"零": 0, "一": 1, "二": 2, "两": 2, "三": 3, "四": 4, "五": 5,
               "六": 6, "七": 7, "八": 8, "九": 9, "十": 10}
    if raw in mapping:
        return mapping[raw]
    if "十" in raw and len(raw) <= 3:
        left, _, right = raw.partition("十")
        return mapping.get(left, 1) * 10 + mapping.get(right, 0)
    return None


def _shift_years(day: date, years: int) -> date:
    try:
        return day.replace(year=day.year - years)
    except ValueError:  # February 29.
        return day.replace(year=day.year - years, day=28)


def _extract_requested_date_range(request_text: str) -> Optional[tuple[str, str]]:
    """Return an explicit natural-language date constraint, independent of the planner."""
    text = str(request_text or "")
    today = datetime.now(timezone(timedelta(hours=8))).date()
    explicit = re.findall(r"(?<!\d)(20\d{2})[-/.\u5e74](\d{1,2})[-/.\u6708](\d{1,2})(?:\u65e5)?(?!\d)", text)
    if len(explicit) >= 2:
        dates = sorted(date(int(year), int(month), int(day)) for year, month, day in explicit[:2])
        return dates[0].isoformat(), dates[-1].isoformat()
    compact = re.findall(r"(?<!\d)(20\d{6})(?!\d)", text)
    if len(compact) >= 2:
        dates = sorted(datetime.strptime(value, "%Y%m%d").date() for value in compact[:2])
        return dates[0].isoformat(), dates[-1].isoformat()
    for unit, multiplier in (("年", 1), ("个月", 30), ("月", 30), ("天", 1)):
        match = re.search(rf"(?:最近|近|过去)\s*([\d零一二两三四五六七八九十]+)\s*{unit}", text)
        if not match:
            continue
        count = _chinese_number(match.group(1))
        if not count or count > 20_000:
            continue
        start = _shift_years(today, count) if unit == "年" else today - timedelta(days=count * multiplier)
        return start.isoformat(), today.isoformat()
    if "今年" in text:
        return date(today.year, 1, 1).isoformat(), today.isoformat()
    return None


class DataAcquisitionPlanner:
    def __init__(self, *, session: Optional[requests.Session] = None):
        config = get_config()
        keys = list(getattr(config, "deepseek_api_keys", None) or [])
        self.api_key = str((keys[0] if keys else "") or "").strip()
        self.base_url = str(os.getenv("DATA_ACQUISITION_LLM_BASE_URL") or "https://api.deepseek.com").rstrip("/")
        self.model = str(
            os.getenv("DATA_ACQUISITION_LLM_MODEL")
            or os.getenv("ESSAY_ANALYSIS_MODEL")
            or DEFAULT_ESSAY_MODEL
        ).strip()
        self.timeout = max(10, min(int(os.getenv("DATA_ACQUISITION_LLM_TIMEOUT_SEC", "120")), 300))
        self.session = session or requests.Session()

    def plan(self, request_text: str) -> Dict[str, Any]:
        if not self.api_key:
            raise DataAcquisitionError("DeepSeek API key is not configured")
        catalog = {
            "tushare": TUSHARE_CATALOG,
            "zsxq": {"research_notes": "知识星球 MCP 增量同步后的本地调研纪要（含图片和附件链接）"},
            "cninfo": {"announcements": "巨潮官网公告现场检索（可下载PDF）"},
            "monitor": {"events": "统一投资情报事件流（仅用于其他情报，不代替公告、新闻或研报官方接口）"},
            "tianyancha": {key: key for key in TYC_COMMANDS},
        }
        payload = {
            "model": self.model,
            "messages": [
                {"role": "system", "content": PLANNER_SYSTEM},
                {"role": "user", "content": json.dumps({
                    "today": datetime.now(timezone(timedelta(hours=8))).date().isoformat(),
                    "request": request_text,
                    "capability_catalog": catalog,
                }, ensure_ascii=False)},
            ],
            "response_format": {"type": "json_object"},
            "thinking": {"type": "disabled"},
            "temperature": 0.1,
            "max_tokens": 6000,
            "stream": False,
        }
        try:
            response = self.session.post(
                f"{self.base_url}/chat/completions",
                headers={"Authorization": f"Bearer {self.api_key}", "Content-Type": "application/json"},
                json=payload, timeout=self.timeout,
            )
            response.raise_for_status()
            body = response.json()
            content = body["choices"][0]["message"]["content"]
            parsed = json.loads(repair_json(content))
        except Exception as exc:
            logger.warning("Data acquisition planning failed: %s", type(exc).__name__)
            raise DataAcquisitionError("大模型暂时无法生成取数计划，请稍后重试") from exc
        return self._normalize(parsed, request_text)

    def _normalize(self, raw: Dict[str, Any], request_text: str) -> Dict[str, Any]:
        tasks = []
        for index, item in enumerate(raw.get("tasks") or []):
            if not isinstance(item, dict):
                continue
            source = str(item.get("source") or "").strip().lower()
            resource = str(item.get("resource") or "").strip().lower()
            if not DataAcquisitionService.is_supported(source, resource):
                continue
            fields = item.get("fields") if isinstance(item.get("fields"), list) else []
            tasks.append({
                "id": f"task-{index + 1}", "source": source, "resource": resource,
                "label": str(item.get("label") or TUSHARE_CATALOG.get(resource) or resource)[:80],
                "reason": str(item.get("reason") or "")[:300],
                "params": item.get("params") if isinstance(item.get("params"), dict) else {},
                "fields": [str(value) for value in fields[:100]],
            })
            if len(tasks) >= MAX_TASKS:
                break
        if not tasks:
            raise DataAcquisitionError("大模型没有生成可执行的数据任务，请补充股票、日期或数据类型")
        scope = DataAcquisitionService.normalize_scope(raw.get("scope"), tasks, request_text)
        return {
            "title": str(raw.get("title") or "一站式财经数据包")[:120],
            "objective": str(raw.get("objective") or request_text)[:1000],
            "tasks": tasks,
            "scope": scope,
            "include_files": bool(raw.get("include_files")) or bool(_FILE_REQUEST_RE.search(request_text)),
            "output_formats": ["json", "csv", "xlsx", "zip"],
            "caveats": [str(value)[:300] for value in (raw.get("caveats") or [])[:10]],
            "model": self.model,
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }


class DataAcquisitionService:
    def __init__(self, *, planner: Optional[DataAcquisitionPlanner] = None,
                 financial: Optional[FinancialDataService] = None,
                 monitor: Optional[InvestmentMonitorService] = None,
                 cninfo: Optional[CninfoAnnouncementService] = None,
                 output_root: Optional[Path] = None):
        self.planner = planner or DataAcquisitionPlanner()
        self.financial = financial or FinancialDataService()
        self.monitor = monitor or InvestmentMonitorService()
        self.cninfo = cninfo or CninfoAnnouncementService()
        config = get_config()
        default_root = Path(config.database_path).expanduser().resolve().parent / "data_acquisition"
        self.output_root = Path(output_root or os.getenv("DATA_ACQUISITION_OUTPUT_DIR") or default_root).resolve()
        self.output_root.mkdir(parents=True, exist_ok=True)

    @staticmethod
    def is_supported(source: str, resource: str) -> bool:
        if source == "tushare":
            return bool(_RESOURCE_RE.fullmatch(resource))
        if source == "zsxq":
            return resource == "research_notes"
        if source == "monitor":
            return resource == "events"
        if source == "cninfo":
            return resource == "announcements"
        return source == "tianyancha" and resource in TYC_COMMANDS

    def capabilities(self) -> Dict[str, Any]:
        return {
            "sources": [
                {"key": "tushare", "name": "Tushare Pro", "mode": "live", "available": self.financial.tushare.available,
                 "resources": [{"key": key, "name": value} for key, value in TUSHARE_CATALOG.items()],
                 "scope": "支持已授权的全部 Tushare api_name"},
                {"key": "zsxq", "name": "知识星球", "mode": "MCP 增量同步 + SQLite", "available": True,
                 "resources": [{"key": "research_notes", "name": "调研纪要、图片与附件索引"}]},
                {"key": "cninfo", "name": "巨潮资讯", "mode": "官方接口现场检索", "available": True,
                 "resources": [{"key": "announcements", "name": "上市公司公告与PDF原文"}]},
                {"key": "monitor", "name": "统一情报", "mode": "SQLite", "available": True,
                 "resources": [{"key": "events", "name": "其他多渠道情报事件"}]},
                {"key": "tianyancha", "name": "天眼查", "mode": "官方 MCP CLI 现场查询", "available": shutil.which("tyc") is not None,
                 "resources": [{"key": key, "name": key.replace("_", " ")} for key in TYC_COMMANDS]},
            ],
            "planner": {"available": bool(self.planner.api_key), "model": self.planner.model,
                        "max_tasks": MAX_TASKS, "max_rows_per_dataset": MAX_ROWS_PER_DATASET},
        }

    def plan(self, request_text: str) -> Dict[str, Any]:
        request_text = str(request_text or "").strip()
        return self._validate_plan(self.planner.plan(request_text), request_text)

    def run(
        self,
        request_text: str,
        plan: Optional[Dict[str, Any]] = None,
        *,
        progress_callback: Optional[Callable[[Dict[str, Any]], None]] = None,
    ) -> Dict[str, Any]:
        normalized_plan = self._validate_plan(plan or self.plan(request_text), request_text)
        self._emit_progress(
            progress_callback,
            progress=3,
            phase="validating",
            message="取数范围与渠道计划已校验",
            completed_tasks=0,
            total_tasks=len(normalized_plan["tasks"]),
        )
        now = datetime.now(timezone.utc)
        digest = hashlib.sha256(f"{now.isoformat()}:{request_text}".encode()).hexdigest()[:10]
        job_id = f"{now.strftime('%Y%m%dT%H%M%SZ')}-{digest}"
        final_dir = self.output_root / job_id
        if final_dir.exists():
            raise DataAcquisitionError("数据任务编号冲突，请重试")
        temp_dir = Path(tempfile.mkdtemp(prefix=f".{job_id}-", dir=self.output_root))
        datasets: List[Dict[str, Any]] = []
        try:
            total_tasks = len(normalized_plan["tasks"])
            for task_index, task in enumerate(normalized_plan["tasks"]):
                self._emit_progress(
                    progress_callback,
                    progress=5 + round((task_index / max(total_tasks, 1)) * 65),
                    phase="fetching",
                    message=f"正在获取：{task['label']}",
                    completed_tasks=task_index,
                    total_tasks=total_tasks,
                    current_task_id=task["id"],
                    current_source=task["source"],
                )
                try:
                    rows = self._execute_task(task, normalized_plan["scope"])
                    datasets.append({"task": task, "status": "success", "rows": rows[:MAX_ROWS_PER_DATASET],
                                     "row_count": len(rows[:MAX_ROWS_PER_DATASET]), "error": None,
                                     "applied_scope": normalized_plan["scope"]})
                except Exception as exc:  # One source must not invalidate an otherwise useful package.
                    logger.warning("Acquisition task failed source=%s resource=%s error=%s",
                                   task["source"], task["resource"], type(exc).__name__)
                    datasets.append({"task": task, "status": "failed", "rows": [], "row_count": 0,
                                     "error": self._safe_error(exc), "applied_scope": normalized_plan["scope"]})
                self._emit_progress(
                    progress_callback,
                    progress=5 + round(((task_index + 1) / max(total_tasks, 1)) * 65),
                    phase="fetching",
                    message=f"已完成 {task_index + 1}/{total_tasks} 个渠道任务",
                    completed_tasks=task_index + 1,
                    total_tasks=total_tasks,
                    current_task_id=task["id"],
                    current_source=task["source"],
                )
            manifest = self._write_artifacts(
                temp_dir,
                job_id,
                request_text,
                normalized_plan,
                datasets,
                progress_callback=progress_callback,
            )
            temp_dir.rename(final_dir)
            self._emit_progress(
                progress_callback,
                progress=100,
                phase="completed",
                message="数据包已生成，可以下载",
                completed_tasks=total_tasks,
                total_tasks=total_tasks,
                job_id=job_id,
            )
            return manifest
        except Exception:
            shutil.rmtree(temp_dir, ignore_errors=True)
            raise

    def list_jobs(self, limit: int = 20) -> Dict[str, Any]:
        jobs = []
        for path in sorted(self.output_root.glob("*/manifest.json"), reverse=True)[:max(1, min(limit, 100))]:
            try:
                jobs.append(json.loads(path.read_text(encoding="utf-8")))
            except (OSError, ValueError):
                continue
        return {"items": jobs, "total": len(jobs)}

    def get_job(self, job_id: str) -> Dict[str, Any]:
        path = self._job_path(job_id) / "manifest.json"
        if not path.is_file():
            raise DataAcquisitionError("数据包不存在")
        return json.loads(path.read_text(encoding="utf-8"))

    def package_path(self, job_id: str) -> Path:
        path = self._job_path(job_id) / f"{job_id}.zip"
        if not path.is_file():
            raise DataAcquisitionError("数据包不存在")
        return path

    def _job_path(self, job_id: str) -> Path:
        if not re.fullmatch(r"[A-Za-z0-9_-]{8,80}", job_id or ""):
            raise DataAcquisitionError("无效的数据包编号")
        path = (self.output_root / job_id).resolve()
        if self.output_root not in path.parents:
            raise DataAcquisitionError("无效的数据包路径")
        return path

    def _validate_plan(self, plan: Dict[str, Any], request_text: str) -> Dict[str, Any]:
        if not isinstance(plan, dict):
            raise DataAcquisitionError("取数计划必须是对象")
        tasks = []
        for index, raw in enumerate((plan.get("tasks") or [])[:MAX_TASKS]):
            if not isinstance(raw, dict):
                continue
            source = str(raw.get("source") or "").strip().lower()
            resource = str(raw.get("resource") or "").strip().lower()
            if not self.is_supported(source, resource):
                raise DataAcquisitionError(f"不支持的数据任务：{source}.{resource}")
            params = raw.get("params") if isinstance(raw.get("params"), dict) else {}
            fields = raw.get("fields") if isinstance(raw.get("fields"), list) else []
            tasks.append({"id": str(raw.get("id") or f"task-{index + 1}")[:50], "source": source,
                          "resource": resource, "label": str(raw.get("label") or resource)[:80],
                          "reason": str(raw.get("reason") or "")[:300], "params": _json_safe(params),
                          "fields": [str(value) for value in fields[:100]]})
        if not tasks:
            raise DataAcquisitionError("取数计划没有可执行任务")
        scope = self.normalize_scope(plan.get("scope"), tasks, request_text)
        self._enforce_research_report_constraints(tasks, scope, request_text)
        return {"title": str(plan.get("title") or "一站式财经数据包")[:120],
                "objective": str(plan.get("objective") or request_text)[:1000], "tasks": tasks,
                "scope": scope,
                "include_files": bool(plan.get("include_files")) or bool(_FILE_REQUEST_RE.search(request_text)),
                "output_formats": ["json", "csv", "xlsx", "zip"],
                "caveats": [str(value)[:300] for value in (plan.get("caveats") or [])[:10]],
                "model": str(plan.get("model") or self.planner.model)[:100],
                "generated_at": str(plan.get("generated_at") or datetime.now(timezone.utc).isoformat())}

    @staticmethod
    def normalize_scope(raw_scope: Any, tasks: List[Dict[str, Any]], request_text: str) -> Dict[str, Any]:
        scope = raw_scope if isinstance(raw_scope, dict) else {}
        serialized = json.dumps([task.get("params") or {} for task in tasks], ensure_ascii=False)
        symbol_values = scope.get("symbols") if isinstance(scope.get("symbols"), list) else []
        symbols = {str(value).strip().upper() for value in symbol_values if str(value).strip()}
        for match in re.finditer(r"(?<!\d)(\d{6})(?:\.(SH|SS|SZ|BJ))?(?!\d)", f"{request_text}\n{serialized}", re.I):
            digits, suffix = match.group(1), (match.group(2) or "").upper().replace("SS", "SH")
            if not suffix:
                suffix = "BJ" if digits.startswith(("4", "8")) else "SH" if digits.startswith(("6", "9")) else "SZ"
            symbols.add(f"{digits}.{suffix}")
        names = set()
        keywords = set()
        for value in scope.get("company_names") if isinstance(scope.get("company_names"), list) else []:
            if str(value).strip():
                names.add(str(value).strip())
        for value in scope.get("keywords") if isinstance(scope.get("keywords"), list) else []:
            keywords.update(_split_search_terms(value))
        for task in tasks:
            params = task.get("params") or {}
            for key in ("company_name", "company_names"):
                values = params.get(key, [])
                values = values if isinstance(values, list) else [values]
                names.update(str(value).strip() for value in values if str(value).strip())
            if task.get("resource") == "research_report":
                for key in ("topics", "keywords", "keyword", "query"):
                    keywords.update(_split_search_terms(params.get(key)))
        request_and_names = f"{request_text}\n{' '.join(names)}".lower()
        for watch_symbol, watch_name in EssayDailyReportService._daily_watchlist():
            if watch_name.lower() in request_and_names:
                names.add(watch_name)
                symbols.add(watch_symbol)
        start_date = str(scope.get("start_date") or "").strip()
        end_date = str(scope.get("end_date") or "").strip()
        if not start_date or not end_date:
            for task in tasks:
                params = task.get("params") or {}
                start_date = start_date or str(params.get("start_date") or params.get("created_from") or "").strip()
                end_date = end_date or str(params.get("end_date") or params.get("created_to") or "").strip()
        return {"symbols": sorted(symbols), "company_names": sorted(names), "keywords": sorted(keywords),
                "start_date": start_date, "end_date": end_date,
                "market_wide": bool(scope.get("market_wide", False))}

    @staticmethod
    def _enforce_research_report_constraints(
        tasks: List[Dict[str, Any]], scope: Dict[str, Any], request_text: str,
    ) -> None:
        """Make report selection deterministic even when the LLM plan is incomplete."""
        report_tasks = [
            task for task in tasks
            if task.get("source") == "tushare" and task.get("resource") == "research_report"
        ]
        if not report_tasks:
            return

        requested_range = _extract_requested_date_range(request_text)
        if requested_range:
            scope["start_date"], scope["end_date"] = requested_range
        if not scope.get("start_date") or not scope.get("end_date"):
            today = datetime.now(timezone(timedelta(hours=8))).date()
            scope["start_date"] = (today - timedelta(days=365)).isoformat()
            scope["end_date"] = today.isoformat()

        topics = list(scope.get("keywords") or [])
        topics.extend(_extract_report_topics(request_text))
        for task in report_tasks:
            params = task.get("params") or {}
            for key in ("topics", "keywords", "keyword", "query"):
                topics.extend(_split_search_terms(params.get(key)))
        topics = list(dict.fromkeys(value for value in topics if value))
        scope["keywords"] = sorted(set(topics))
        if topics:
            # Topic-wide retrieval is still a scoped query; market_wide means no textual filter.
            scope["market_wide"] = False

        has_entity_selector = bool(scope.get("symbols") or scope.get("company_names"))
        if not topics and not has_entity_selector and not scope.get("market_wide"):
            raise DataAcquisitionError("研报任务缺少主题、公司或股票范围，已阻止全量获取")

        strict_depth = bool(re.search(r"(?:只要|仅要|必须|全部|都要).{0,8}深度", request_text))
        prefer_depth = strict_depth or bool(re.search(r"(?:深度研究|深度研报|深度报告|最好.{0,8}深度)", request_text))
        mode = "any" if len(topics) > 1 or re.search(r"(?:或者|或|\bor\b)", request_text, re.I) else "all"
        for task in report_tasks:
            params = dict(task.get("params") or {})
            params.update({
                "start_date": str(scope["start_date"]).replace("-", ""),
                "end_date": str(scope["end_date"]).replace("-", ""),
                "topics": topics,
                "keyword_mode": mode,
                "depth_preference": "strict" if strict_depth else "prefer" if prefer_depth else "none",
                "ai_filter": bool(params.get("ai_filter", True)) if topics else False,
                "max_results": DataAcquisitionService._bounded_report_result_limit(params.get("max_results")),
            })
            task["params"] = params

    @staticmethod
    def _bounded_report_result_limit(value: Any) -> int:
        try:
            return max(1, min(int(value or MAX_RESEARCH_REPORT_RESULTS), MAX_ROWS_PER_DATASET))
        except (TypeError, ValueError):
            return MAX_RESEARCH_REPORT_RESULTS

    def _execute_task(self, task: Dict[str, Any], scope: Dict[str, Any]) -> List[Dict[str, Any]]:
        source, resource, params = task["source"], task["resource"], dict(task["params"])
        if source == "tushare":
            return self._execute_tushare(task, scope)
        if source == "zsxq":
            created_from = params.pop("start_date", params.get("created_from", None))
            created_to = params.pop("end_date", params.get("created_to", None))
            keywords = params.pop("keywords", None)
            queries = keywords if isinstance(keywords, list) else [params.pop("query", "")]
            symbols = params.pop("symbols", None)
            if symbols:
                values = symbols if isinstance(symbols, list) else [symbols]
                queries.extend(values)
            scoped_queries = [
                *(scope.get("company_names") or []),
                *(scope.get("symbols") or []),
                *(scope.get("keywords") or []),
            ]
            queries = self._expand_zsxq_queries([*queries, *scoped_queries])
            rows_by_id: Dict[str, Dict[str, Any]] = {}
            for value in queries or [""]:
                current = {key: item for key, item in params.items() if key in {
                    "group_id", "digested", "page", "page_size",
                }}
                current["page_size"] = min(int(current.get("page_size") or 100), 100)
                if created_from:
                    current["created_from"] = created_from
                if created_to:
                    current["created_to"] = created_to
                text = str(value or "").strip()
                if text:
                    current["symbol" if re.fullmatch(r"\d{6}(?:\.(?:SH|SZ|BJ))?", text, re.I) else "query"] = text
                result = self.financial.query(source=source, resource=resource, params=current, fields=None)
                for row in result.get("rows") or []:
                    rows_by_id[str(row.get("topic_id") or len(rows_by_id))] = row
            return self._filter_rows([_json_safe(row) for row in rows_by_id.values()], source, resource, scope)
        if source == "cninfo":
            start = self._as_date(params.get("start_date") or scope.get("start_date"), default_days=30)
            end = self._as_date(params.get("end_date") or scope.get("end_date"), default_days=0)
            raw_categories = params.get("categories") or params.get("category") or []
            categories = raw_categories if isinstance(raw_categories, list) else [value for value in str(raw_categories).split(";") if value]
            raw_symbols = params.get("symbols") or params.get("ts_codes") or params.get("ts_code") or params.get("symbol") or []
            symbols = raw_symbols if isinstance(raw_symbols, list) else [value.strip() for value in str(raw_symbols).split(",") if value.strip()]
            symbols = symbols or list(scope.get("symbols") or [])
            if not symbols and not scope.get("market_wide"):
                raise DataAcquisitionError("巨潮现场检索缺少股票代码，已阻止全市场请求")
            rows = self.cninfo.fetch(
                start_date=start, end_date=end, symbols=symbols, categories=categories,
                keyword=str(params.get("keyword") or params.get("query") or ""),
                page_size=min(int(params.get("page_size") or 100), 100),
                max_pages=min(int(params.get("max_pages") or 20), 100),
            )
            return [_json_safe(row) for row in rows]
        if source == "monitor":
            base_params = dict(params)
            raw_symbol = base_params.pop("ts_code", None) or base_params.pop("stock_code", None)
            if raw_symbol and not base_params.get("symbol"):
                base_params["symbol"] = raw_symbol
            raw_keywords = base_params.pop("keywords", None)
            if raw_keywords and not base_params.get("query"):
                values = raw_keywords if isinstance(raw_keywords, list) else [raw_keywords]
                base_params["query"] = " ".join(str(value).strip() for value in values if str(value).strip())
            for key in ("start_date", "end_date"):
                value = str(base_params.get(key) or "")
                if re.fullmatch(r"\d{8}", value):
                    base_params[key] = f"{value[:4]}-{value[4:6]}-{value[6:]}"
            allowed = ({
                "days", "symbol", "perspective", "event_type", "source_key", "query",
                "min_importance", "channel", "evidence_level", "page", "page_size",
            })
            base_params = {key: value for key, value in base_params.items() if key in allowed}
            base_params["page"] = 1
            base_params["page_size"] = min(int(base_params.get("page_size") or 200), 200)
            selectors = list(scope.get("symbols") or []) or list(scope.get("company_names") or [])
            if not selectors and not scope.get("market_wide") and not base_params.get("query") and not base_params.get("symbol"):
                raise DataAcquisitionError("该公告/情报任务没有明确股票或关键词，已阻止全量获取")
            param_sets = []
            for selector in selectors:
                current = dict(base_params)
                if re.fullmatch(r"\d{6}(?:\.(?:SH|SZ|BJ))?", str(selector), re.I):
                    current["symbol"] = selector
                else:
                    current["query"] = selector
                param_sets.append(current)
            if not param_sets:
                param_sets = [base_params]
            rows_by_id: Dict[str, Dict[str, Any]] = {}
            for current in param_sets:
                result = self.monitor.list_events(**current)
                for row in result.get("items") or []:
                    rows_by_id[str(row.get("id") or row.get("external_id") or len(rows_by_id))] = row
            return self._filter_rows([_json_safe(row) for row in rows_by_id.values()], source, resource, scope)
        if source == "tianyancha":
            raw_keys = params.get("search_key") or params.get("searchKey") or params.get("company_name") or ""
            search_keys = raw_keys if isinstance(raw_keys, list) else [raw_keys]
            search_keys = [str(value).strip() for value in search_keys if str(value).strip()]
            search_keys = search_keys or list(scope.get("company_names") or [])
            if not search_keys or any(len(value) > 200 for value in search_keys):
                raise DataAcquisitionError("天眼查任务缺少 company_name/search_key")
            rows = []
            for search_key in search_keys[:20]:
                resolved = self._resolve_tyc_company(search_key)
                resolved_key = str(resolved.get("name") or resolved.get("creditCode") or search_key)
                rows.append({"query_company": search_key, "facet": "entity_resolution", "status": "success", "payload": resolved})
                commands = TYC_FULL_COMMANDS if resource == "company_full" else {resource: TYC_COMMANDS[resource]}
                worker_limit = max(1, min(int(os.getenv("DATA_ACQUISITION_MAX_WORKERS", "6")), 6))
                with ThreadPoolExecutor(max_workers=min(worker_limit, len(commands))) as pool:
                    futures = {pool.submit(self._run_tyc, command, resolved_key): facet for facet, command in commands.items()}
                    for future in as_completed(futures):
                        facet = futures[future]
                        try:
                            payload = future.result()
                            rows.append({"query_company": search_key, "facet": facet, "status": "success", "payload": _json_safe(payload)})
                        except Exception as exc:
                            rows.append({"query_company": search_key, "facet": facet, "status": "failed", "error": self._safe_error(exc)})
            return rows
        raise DataAcquisitionError(f"不支持的数据源：{source}")

    def _execute_tushare(self, task: Dict[str, Any], scope: Dict[str, Any]) -> List[Dict[str, Any]]:
        resource = task["resource"]
        if resource == "research_report":
            return self._execute_research_reports(task, scope)
        query_params = dict(task["params"])
        raw_codes = query_params.pop("ts_codes", query_params.get("ts_code", ""))
        codes = raw_codes if isinstance(raw_codes, list) else [value.strip() for value in str(raw_codes).split(",") if value.strip()]
        codes = codes or list(scope.get("symbols") or [])
        fields = list(task.get("fields") or [])
        if resource in {"news", "major_news"}:
            query_params.pop("ts_code", None)
            for key in ("keyword", "keywords", "query", "company_name", "company_names", "symbol", "symbols"):
                query_params.pop(key, None)
            if resource == "news":
                query_params.setdefault("src", "cls")
            else:
                query_params.setdefault("src", "新浪财经")
            for key, suffix in (("start_date", "00:00:00"), ("end_date", "23:59:59")):
                raw = str(query_params.get(key) or scope.get(key) or "")
                if re.fullmatch(r"\d{8}", raw):
                    raw = f"{raw[:4]}-{raw[4:6]}-{raw[6:]} {suffix}"
                elif re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw):
                    raw = f"{raw} {suffix}"
                if raw:
                    query_params[key] = raw
            codes = []

        param_sets = []
        if codes and resource not in {"news", "major_news"}:
            for code in codes[:20]:
                current = dict(query_params)
                current["ts_code"] = code
                param_sets.append(current)
        else:
            param_sets.append(query_params)

        page_size = 1500 if resource == "news" else 800 if resource == "major_news" else 0
        rows: List[Dict[str, Any]] = []
        for base in param_sets:
            if not page_size:
                result = self.financial.query(source="tushare", resource=resource, params=base, fields=fields or None)
                rows.extend(result.get("rows") or [])
                continue
            for offset in range(0, MAX_ROWS_PER_DATASET, page_size):
                current = {**base, "limit": page_size, "offset": offset}
                result = self.financial.query(source="tushare", resource=resource, params=current, fields=fields or None)
                page = result.get("rows") or []
                rows.extend(page)
                if len(page) < page_size:
                    break
        filtered = self._filter_rows([_json_safe(row) for row in rows], "tushare", resource, scope)
        unique: Dict[str, Dict[str, Any]] = {}
        for row in filtered:
            identity = json.dumps(
                [row.get("ts_code"), row.get("trade_date") or row.get("datetime"), row.get("title"), row.get("url")],
                ensure_ascii=False, default=str,
            )
            unique[identity] = row
        return list(unique.values())

    def _execute_research_reports(self, task: Dict[str, Any], scope: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Recall the full requested period, then rank/filter before exporting anything."""
        params = dict(task.get("params") or {})
        topics: List[str] = []
        for key in ("topics", "keywords", "keyword", "query"):
            topics.extend(_split_search_terms(params.get(key)))
        topics.extend(str(value).strip() for value in scope.get("keywords") or [] if str(value).strip())
        topics = list(dict.fromkeys(topics))
        mode = "all" if str(params.get("keyword_mode") or "any").lower() == "all" else "any"
        depth_preference = str(params.get("depth_preference") or "none").lower()
        ai_filter = bool(params.get("ai_filter")) and bool(topics)
        max_results = self._bounded_report_result_limit(params.get("max_results"))

        start = self._as_date(params.get("start_date") or scope.get("start_date"), default_days=365)
        end = self._as_date(params.get("end_date") or scope.get("end_date"), default_days=0)
        if end < start:
            raise DataAcquisitionError("研报开始日期不能晚于结束日期")
        codes = params.get("ts_codes") or params.get("ts_code") or scope.get("symbols") or []
        codes = codes if isinstance(codes, list) else [value.strip() for value in str(codes).split(",") if value.strip()]
        if not topics and not codes and not scope.get("company_names") and not scope.get("market_wide"):
            raise DataAcquisitionError("研报筛选条件丢失，已阻止全量下载")

        fields = list(dict.fromkeys([
            *(task.get("fields") or []), "trade_date", "abstr", "title", "report_type",
            "author", "name", "ts_code", "inst_csname", "ind_name", "url",
        ]))
        upstream_base = {
            key: value for key, value in params.items()
            if key not in _REPORT_INTERNAL_PARAMS and key not in {"limit", "offset", "ts_codes"}
            and value not in (None, "", [])
        }
        upstream_base.pop("start_date", None)
        upstream_base.pop("end_date", None)

        selectors: List[Optional[str]] = [str(value) for value in codes[:20]] if codes else [None]
        if codes and topics:
            # Industry/strategy reports often have no ts_code and must join the topic recall.
            selectors.append(None)
        rows_by_identity: Dict[str, Dict[str, Any]] = {}
        scanned = 0
        recall_truncated = False
        for selector in selectors:
            for window_start, window_end in self._report_date_windows(start, end):
                for offset in range(0, MAX_RESEARCH_REPORT_RECALL_ROWS, 1000):
                    current = {
                        **upstream_base,
                        "start_date": window_start.strftime("%Y%m%d"),
                        "end_date": window_end.strftime("%Y%m%d"),
                        "limit": 1000,
                        "offset": offset,
                    }
                    if selector:
                        current["ts_code"] = selector
                    result = self.financial.query(
                        source="tushare", resource="research_report", params=current, fields=fields,
                    )
                    page = result.get("rows") or []
                    scanned += len(page)
                    for raw_row in page:
                        row = _json_safe(raw_row)
                        scored = self._score_research_report(row, topics, mode)
                        if scored is None:
                            continue
                        if depth_preference == "strict" and int(scored.get("深度评分") or 0) < 45:
                            continue
                        identity = json.dumps(
                            [scored.get("trade_date"), scored.get("title"), scored.get("url")],
                            ensure_ascii=False, default=str,
                        )
                        rows_by_identity[identity] = scored
                    if scanned >= MAX_RESEARCH_REPORT_RECALL_ROWS:
                        recall_truncated = True
                        break
                    if len(page) < 1000:
                        break
                if recall_truncated:
                    break
            if recall_truncated:
                break

        candidates = list(rows_by_identity.values())
        if not topics:
            candidates = self._filter_rows(candidates, "tushare", "research_report", scope)
        candidates.sort(
            key=lambda row: (
                int(row.get("深度评分") or 0) if depth_preference in {"prefer", "strict"} else 0,
                int(row.get("相关性评分") or 0),
                str(row.get("trade_date") or ""),
            ),
            reverse=True,
        )
        review_pool = candidates[:max(max_results * 3, max_results)]
        if ai_filter and review_pool:
            review_pool = self._ai_review_research_reports(review_pool, topics, depth_preference)
        selected = review_pool[:max_results]
        for row in selected:
            row["检索开始日期"] = start.isoformat()
            row["检索结束日期"] = end.isoformat()
            row["候选扫描数"] = scanned
            row["召回是否截断"] = recall_truncated
        return selected

    @staticmethod
    def _report_date_windows(start: date, end: date) -> List[tuple[date, date]]:
        windows: List[tuple[date, date]] = []
        cursor = end
        while cursor >= start:
            window_start = max(start, cursor - timedelta(days=30))
            windows.append((window_start, cursor))
            cursor = window_start - timedelta(days=1)
        return windows

    @staticmethod
    def _score_research_report(
        row: Dict[str, Any], topics: List[str], mode: str,
    ) -> Optional[Dict[str, Any]]:
        title = str(row.get("title") or "")
        abstract = str(row.get("abstr") or "")
        industry = str(row.get("ind_name") or "")
        company = str(row.get("name") or "")
        title_lower, abstract_lower = title.lower(), abstract.lower()
        context_lower = f"{industry} {company}".lower()
        matched_topics: List[str] = []
        matched_terms: List[str] = []
        relevance = 0
        for topic in topics:
            aliases = _REPORT_TOPIC_ALIASES.get(topic, [topic])
            topic_hit = False
            topic_score = 0
            for alias in aliases:
                needle = alias.lower()
                if not needle:
                    continue
                if needle in title_lower:
                    topic_score = max(topic_score, 55 if alias == topic else 42)
                    topic_hit = True
                    matched_terms.append(alias)
                if needle in abstract_lower:
                    topic_score = max(topic_score, 35 if alias == topic else 24)
                    topic_hit = True
                    matched_terms.append(alias)
                if needle in context_lower:
                    topic_score = max(topic_score, 30)
                    topic_hit = True
                    matched_terms.append(alias)
            if topic_hit:
                matched_topics.append(topic)
                relevance += topic_score
        if topics and ((mode == "all" and len(matched_topics) != len(topics)) or (mode != "all" and not matched_topics)):
            return None
        relevance = min(100, relevance + max(0, len(set(matched_topics)) - 1) * 10)

        report_type = str(row.get("report_type") or "")
        depth_text = f"{title} {report_type}"
        depth = 0
        depth_reasons: List[str] = []
        for term, points in (("深度", 55), ("专题", 45), ("产业链", 35), ("全景", 35),
                             ("白皮书", 35), ("年度策略", 30), ("行业研究", 25)):
            if term in depth_text:
                depth = max(depth, points)
                depth_reasons.append(term)
        if len(abstract) >= 500:
            depth += 20
            depth_reasons.append("长摘要")
        elif len(abstract) >= 200:
            depth += 10
            depth_reasons.append("完整摘要")
        if row.get("url"):
            depth += 10
            depth_reasons.append("可下载PDF")
        if re.search(r"(?:日报|周报|早报|晨报|快评|点评)", title):
            depth -= 30
            depth_reasons.append("短评类扣分")
        enriched = dict(row)
        enriched.update({
            "筛选命中主题": "、".join(dict.fromkeys(matched_topics)),
            "筛选命中词": "、".join(dict.fromkeys(matched_terms)),
            "相关性评分": relevance,
            "深度评分": max(0, min(100, depth)),
            "深度判定": "高" if depth >= 60 else "中" if depth >= 35 else "普通",
            "深度判定依据": "、".join(dict.fromkeys(depth_reasons)) or "未出现深度研究特征",
            "AI语义复核": "待复核",
            "AI复核说明": "",
        })
        return enriched

    def _ai_review_research_reports(
        self, rows: List[Dict[str, Any]], topics: List[str], depth_preference: str,
    ) -> List[Dict[str, Any]]:
        """Use structured semantic classification; never execute model-generated code."""
        enabled = str(os.getenv("DATA_ACQUISITION_REPORT_AI_FILTER", "1")).lower() not in {"0", "false", "off"}
        api_key = str(getattr(self.planner, "api_key", "") or "").strip()
        base_url = str(getattr(self.planner, "base_url", "") or "").rstrip("/")
        model = str(getattr(self.planner, "model", "") or DEFAULT_ESSAY_MODEL)
        if not enabled or not api_key or not base_url:
            for row in rows:
                row["AI语义复核"] = "未执行（使用确定性筛选）"
            return rows

        batches = [rows[index:index + 25] for index in range(0, len(rows), 25)]

        def review(batch_index: int, batch: List[Dict[str, Any]]) -> tuple[int, Optional[Dict[int, Dict[str, Any]]]]:
            items = [{
                "id": index,
                "title": str(row.get("title") or "")[:300],
                "abstract": str(row.get("abstr") or "")[:1400],
                "industry": str(row.get("ind_name") or "")[:100],
                "report_type": str(row.get("report_type") or "")[:100],
            } for index, row in enumerate(batch)]
            system = (
                "你是券商研报语义筛选器。用户主题之间是 OR；只根据标题、摘要和行业判断。"
                "概念真正相关才 relevant=true，偶然出现或仅宏观提及应为false。"
                "depth_score评估是否是系统深度研究，而非日报/周报/事件快评。"
                "输出严格JSON：{\"items\":[{\"id\":0,\"relevant\":true,\"score\":0-100,"
                "\"depth_score\":0-100,\"matched_topics\":[],\"reason\":\"20字内\"}]}。"
            )
            payload = {
                "model": model,
                "messages": [
                    {"role": "system", "content": system},
                    {"role": "user", "content": json.dumps({"topics": topics, "reports": items}, ensure_ascii=False)},
                ],
                "response_format": {"type": "json_object"},
                "thinking": {"type": "disabled"},
                "temperature": 0,
                "max_tokens": 3500,
                "stream": False,
            }
            try:
                response = requests.post(
                    f"{base_url}/chat/completions",
                    headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                    json=payload,
                    timeout=max(20, min(int(getattr(self.planner, "timeout", 120)), 180)),
                )
                response.raise_for_status()
                parsed = json.loads(repair_json(response.json()["choices"][0]["message"]["content"]))
                reviewed = {
                    int(item["id"]): item for item in parsed.get("items") or []
                    if isinstance(item, dict) and str(item.get("id", "")).isdigit()
                }
                return batch_index, reviewed
            except Exception:
                logger.warning("Research report AI review batch failed index=%s", batch_index)
                return batch_index, None

        reviews: Dict[int, Optional[Dict[int, Dict[str, Any]]]] = {}
        workers = min(6, len(batches))
        with ThreadPoolExecutor(max_workers=max(1, workers)) as pool:
            futures = [pool.submit(review, index, batch) for index, batch in enumerate(batches)]
            for future in as_completed(futures):
                batch_index, result = future.result()
                reviews[batch_index] = result

        kept: List[Dict[str, Any]] = []
        for batch_index, batch in enumerate(batches):
            reviewed = reviews.get(batch_index)
            for local_index, row in enumerate(batch):
                item = reviewed.get(local_index) if reviewed else None
                if not item:
                    row["AI语义复核"] = "上游不可用，已回退确定性筛选"
                    kept.append(row)
                    continue
                ai_score = max(0, min(int(item.get("score") or 0), 100))
                ai_depth = max(0, min(int(item.get("depth_score") or 0), 100))
                row["AI语义复核"] = "通过" if item.get("relevant") else "排除"
                row["AI相关性评分"] = ai_score
                row["AI深度评分"] = ai_depth
                row["AI复核说明"] = str(item.get("reason") or "")[:120]
                relevant = bool(item.get("relevant")) and ai_score >= 50
                deep_enough = depth_preference != "strict" or max(ai_depth, int(row.get("深度评分") or 0)) >= 45
                if relevant and deep_enough:
                    kept.append(row)
        kept.sort(
            key=lambda row: (
                int(row.get("AI深度评分") or row.get("深度评分") or 0),
                int(row.get("AI相关性评分") or row.get("相关性评分") or 0),
                str(row.get("trade_date") or ""),
            ), reverse=True,
        )
        return kept

    @staticmethod
    def _run_tyc(command: List[str], search_key: str) -> Any:
        completed = subprocess.run(["tyc", "--compact", *command, search_key], capture_output=True, text=True, timeout=90, check=False)
        if completed.returncode != 0:
            raise DataAcquisitionError("天眼查查询失败，请检查该维度权限或企业名称")
        try:
            return json.loads(completed.stdout)
        except ValueError as exc:
            raise DataAcquisitionError("天眼查返回格式异常") from exc

    def _resolve_tyc_company(self, search_key: str) -> Dict[str, Any]:
        payload = self._run_tyc(["company", "companies"], search_key)
        items = payload.get("items") if isinstance(payload, dict) else None
        if not isinstance(items, list) or not items:
            raise DataAcquisitionError(f"天眼查没有找到企业：{search_key}")
        candidates = [item for item in items if isinstance(item, dict)]
        # L0 search is relevance-ranked by Tianyancha; re-sorting abbreviated listed-company
        # names locally can select an unrelated exact-substring company.
        return _json_safe(candidates[0])

    @staticmethod
    def _as_date(value: Any, *, default_days: int) -> date:
        raw = str(value or "").strip()
        if not raw:
            return datetime.now(timezone(timedelta(hours=8))).date() - timedelta(days=default_days)
        try:
            return datetime.strptime(raw.replace("-", "")[:8], "%Y%m%d").date()
        except ValueError as exc:
            raise DataAcquisitionError(f"无效日期：{raw}") from exc

    def _download_files(self, directory: Path, task: Dict[str, Any], rows: List[Dict[str, Any]], stem: str) -> List[Dict[str, Any]]:
        candidates: List[Dict[str, str]] = []
        if task["source"] == "cninfo":
            candidates.extend({"url": str(row.get("pdf_url") or ""), "name": str(row.get("title") or row.get("announcement_id") or "公告") + ".pdf"} for row in rows)
        elif task["source"] == "tushare" and task["resource"] == "research_report":
            candidates.extend({"url": str(row.get("url") or ""), "name": str(row.get("title") or "研报") + ".pdf"} for row in rows)
        elif task["source"] == "zsxq":
            from src.services.zsxq_mcp_sync_service import ZsxqMcpSyncService
            resolver = ZsxqMcpSyncService()
            for row in rows:
                topic_id = str(row.get("topic_id") or "")
                for kind, id_key in (("files", "file_id"), ("images", "image_id")):
                    for asset in row.get(kind) or []:
                        asset_id = str(asset.get(id_key) or "")
                        if not topic_id or not asset_id:
                            continue
                        try:
                            url = resolver.resolve_media_url_sync(topic_id, kind, asset_id)
                            candidates.append({"url": url, "name": str(asset.get("name") or asset.get("filename") or f"{topic_id}_{asset_id}")})
                        except Exception as exc:
                            candidates.append({"url": "", "name": str(asset.get("name") or asset_id), "error": self._safe_error(exc)})
        elif task["source"] == "tianyancha":
            for row in rows:
                self._collect_pdf_urls(row.get("payload"), candidates)

        target_dir = directory / "attachments" / task["source"] / stem
        results: List[Dict[str, Any]] = []
        seen = set()
        total_bytes = 0
        for index, candidate in enumerate(candidates[:MAX_DOWNLOAD_FILES], start=1):
            url = str(candidate.get("url") or "").strip()
            name = str(candidate.get("name") or f"file_{index}")
            identity = url or f"error:{name}"
            if identity in seen:
                continue
            seen.add(identity)
            if candidate.get("error") or not url:
                results.append({"name": name, "status": "failed", "error": candidate.get("error") or "原始文件没有可下载地址"})
                continue
            try:
                parsed = urlparse(url)
                if parsed.scheme != "https":
                    raise DataAcquisitionError("只允许下载 HTTPS 原始文件")
                if task["source"] == "cninfo" and parsed.hostname != "static.cninfo.com.cn":
                    raise DataAcquisitionError("巨潮文件域名不受信任")
                target_dir.mkdir(parents=True, exist_ok=True)
                safe_name = self._safe_filename(Path(unquote(parsed.path)).name or name)
                if "." not in safe_name:
                    safe_name += ".pdf" if task["resource"] in {"announcements", "research_report"} else ".bin"
                target = target_dir / f"{index:03d}_{safe_name}"
                size = self._download_one(url, target, remaining=max(0, MAX_PACKAGE_FILE_BYTES - total_bytes))
                total_bytes += size
                results.append({"name": name, "status": "success", "path": str(target.relative_to(directory)), "size_bytes": size})
            except Exception as exc:
                results.append({"name": name, "status": "failed", "error": self._safe_error(exc)})
        if len(candidates) > MAX_DOWNLOAD_FILES:
            results.append({"name": "download_limit", "status": "failed", "error": f"原始文件超过单包上限 {MAX_DOWNLOAD_FILES} 个"})
        return results

    @staticmethod
    def _collect_pdf_urls(value: Any, candidates: List[Dict[str, str]]) -> None:
        if isinstance(value, dict):
            for key, item in value.items():
                if isinstance(item, str) and item.startswith("https://") and ("pdf" in key.lower() or ".pdf" in item.lower()):
                    candidates.append({"url": item, "name": Path(urlparse(item).path).name or "天眼查原始文件.pdf"})
                else:
                    DataAcquisitionService._collect_pdf_urls(item, candidates)
        elif isinstance(value, list):
            for item in value:
                DataAcquisitionService._collect_pdf_urls(item, candidates)

    @staticmethod
    def _download_one(url: str, target: Path, *, remaining: int) -> int:
        if remaining <= 0:
            raise DataAcquisitionError("原始文件达到单包 500MB 上限")
        limit = min(MAX_FILE_BYTES, remaining)
        temporary = target.with_suffix(target.suffix + ".part")
        session = requests.Session()
        response: Optional[requests.Response] = None
        try:
            hostname = (urlparse(url).hostname or "").lower()
            headers = {
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/140 Safari/537.36",
                "Referer": "https://data.eastmoney.com/report/" if hostname == "pdf.dfcfw.com" else "https://www.cninfo.com.cn/",
            }
            response = session.get(url, headers=headers, timeout=(10, 90), stream=True)
            if hostname == "pdf.dfcfw.com" and response.status_code == 200:
                declared = int(response.headers.get("Content-Length") or 0)
                if 0 < declared < 16 * 1024:
                    challenge = response.content.decode("utf-8", errors="ignore")
                    cookies = DataAcquisitionService._solve_eastmoney_pdf_challenge(challenge)
                    response.close()
                    response = None
                    if not cookies:
                        raise DataAcquisitionError("研报 PDF 下载验证失败")
                    for key, value in cookies.items():
                        session.cookies.set(key, value, domain="pdf.dfcfw.com", path="/")
                    response = session.get(url, headers=headers, timeout=(10, 90), stream=True)
            with response:
                response.raise_for_status()
                declared = int(response.headers.get("Content-Length") or 0)
                if declared > limit:
                    raise DataAcquisitionError("原始文件超过大小上限")
                size = 0
                first_chunk = True
                with temporary.open("wb") as output:
                    for chunk in response.iter_content(64 * 1024):
                        if not chunk:
                            continue
                        if first_chunk and ".pdf" in urlparse(url).path.lower() and not chunk.startswith(b"%PDF-"):
                            raise DataAcquisitionError("上游未返回有效 PDF 文件")
                        first_chunk = False
                        size += len(chunk)
                        if size > limit:
                            raise DataAcquisitionError("原始文件超过大小上限")
                        output.write(chunk)
            temporary.replace(target)
            return size
        finally:
            if response is not None:
                response.close()
            session.close()
            temporary.unlink(missing_ok=True)

    @staticmethod
    def _solve_eastmoney_pdf_challenge(script: str) -> Dict[str, str]:
        if "EO_Bot_Ssid" not in script or "__tst_status" not in script:
            return {}
        ssid_match = re.search(r't=a\[_0x[0-9a-f]+\("0x7"\)\]\(t,(\d+)\)', script, re.I)
        block_match = re.search(r"var e=\{(.+?)\},t=0", script, re.S)
        if not ssid_match or not block_match:
            return {}
        values = [int(value) for value in re.findall(r":(\d+)", block_match.group(1))[:3]]
        if len(values) != 3:
            return {}
        return {"EO_Bot_Ssid": ssid_match.group(1), "__tst_status": f"{sum(values)}#"}

    def _write_artifacts(self, directory: Path, job_id: str, request_text: str,
                         plan: Dict[str, Any], datasets: List[Dict[str, Any]], *,
                         progress_callback: Optional[Callable[[Dict[str, Any]], None]] = None) -> Dict[str, Any]:
        generated_at = datetime.now(timezone.utc).isoformat()
        dataset_manifest = []
        workbooks: Dict[str, Workbook] = {}
        self._emit_progress(
            progress_callback,
            progress=72,
            phase="exporting",
            message="正在按渠道生成 JSON、CSV 与 Excel",
            completed_tasks=len(datasets),
            total_tasks=len(datasets),
        )
        for index, dataset in enumerate(datasets, start=1):
            task, rows = dataset["task"], dataset["rows"]
            stem = f"{index:02d}_{self._safe_filename(task['label'])}"
            channel_dir = directory / task["source"]
            channel_dir.mkdir(parents=True, exist_ok=True)
            json_name, csv_name = f"{task['source']}/{stem}.json", f"{task['source']}/{stem}.csv"
            (directory / json_name).write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
            self._write_csv(directory / csv_name, rows)
            workbook = workbooks.setdefault(task["source"], Workbook())
            if workbook.active.title == "Sheet" and len(workbook.sheetnames) == 1:
                workbook.remove(workbook.active)
            self._write_sheet(workbook, stem[:31], rows, dataset.get("error"))
            attachments = self._download_files(directory, task, rows, stem) if plan.get("include_files") else []
            dataset_manifest.append({"task_id": task["id"], "label": task["label"], "source": task["source"],
                                     "resource": task["resource"], "status": dataset["status"],
                                     "row_count": dataset["row_count"], "error": dataset["error"],
                                     "applied_scope": dataset["applied_scope"], "files": [json_name, csv_name],
                                     "attachments": attachments})
            self._emit_progress(
                progress_callback,
                progress=72 + round((index / max(len(datasets), 1)) * 20),
                phase="exporting",
                message=f"已导出 {index}/{len(datasets)} 个数据集",
                completed_tasks=len(datasets),
                total_tasks=len(datasets),
                current_task_id=task["id"],
                current_source=task["source"],
            )
        for source, workbook in workbooks.items():
            workbook.save(directory / source / f"{source}.xlsx")
        success_count = sum(item["status"] == "success" for item in dataset_manifest)
        downloaded_count = sum(sum(file["status"] == "success" for file in item["attachments"]) for item in dataset_manifest)
        failed_download_count = sum(sum(file["status"] == "failed" for file in item["attachments"]) for item in dataset_manifest)
        manifest = {"job_id": job_id, "contract_version": "channel-scoped-v3-live",
                    "title": plan["title"], "request": request_text,
                    "status": "success" if success_count == len(datasets) else ("partial" if success_count else "failed"),
                    "generated_at": generated_at, "plan": plan, "datasets": dataset_manifest,
                    "summary": {"task_count": len(datasets), "success_count": success_count,
                                "failed_count": len(datasets) - success_count,
                                "row_count": sum(item["row_count"] for item in dataset_manifest),
                                "include_files": bool(plan.get("include_files")),
                                "downloaded_file_count": downloaded_count,
                                "failed_file_count": failed_download_count},
                    "download_url": f"/api/v1/data-acquisition/jobs/{job_id}/download",
                    "formats": ["json", "csv", "xlsx", "zip"]}
        (directory / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
        readme = [f"# {plan['title']}", "", f"原始需求：{request_text}", f"生成时间：{generated_at}", "",
                  "数据按渠道分目录；每个数据集同时提供 JSON 与 CSV，每个渠道单独提供 XLSX；manifest.json 记录现场接口、范围、参数、状态和行数。",
                  f"原始文件下载：{'已启用' if plan.get('include_files') else '未请求'}；已下载 {downloaded_count} 个，失败 {failed_download_count} 个。",
                  "失败的数据源不会阻断其他任务，具体原因请查看 manifest.json。"]
        (directory / "README.md").write_text("\n".join(readme), encoding="utf-8")
        zip_name = f"{job_id}.zip"
        self._emit_progress(
            progress_callback,
            progress=96,
            phase="packaging",
            message="正在压缩最终 ZIP 数据包",
            completed_tasks=len(datasets),
            total_tasks=len(datasets),
        )
        with zipfile.ZipFile(directory / zip_name, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for path in sorted(directory.rglob("*")):
                if path.is_file() and path.name != zip_name:
                    archive.write(path, path.relative_to(directory))
        self._emit_progress(
            progress_callback,
            progress=99,
            phase="finalizing",
            message="正在写入来源清单并完成校验",
            completed_tasks=len(datasets),
            total_tasks=len(datasets),
        )
        return manifest

    @staticmethod
    def _emit_progress(
        callback: Optional[Callable[[Dict[str, Any]], None]],
        **update: Any,
    ) -> None:
        if callback is None:
            return
        try:
            callback(update)
        except Exception:  # Progress reporting must never invalidate a useful package.
            logger.debug("Data acquisition progress callback failed", exc_info=True)

    @staticmethod
    def _filter_rows(rows: List[Dict[str, Any]], source: str, resource: str,
                     scope: Dict[str, Any]) -> List[Dict[str, Any]]:
        if scope.get("market_wide"):
            return rows
        symbols = {str(value).upper() for value in scope.get("symbols") or []}
        code_digits = {value.split(".")[0] for value in symbols}
        terms = [str(value).strip() for value in [*(scope.get("company_names") or []),
                                                  *(scope.get("keywords") or [])] if str(value).strip()]
        text_resources = {"news", "major_news", "research_report", "report_rc", "anns_d",
                          "research_notes", "events", "announcements"}
        filtered = []
        for row in rows:
            row_codes = set()
            for key in ("ts_code", "symbol", "stock_code", "code"):
                value = row.get(key)
                if value:
                    row_codes.add(str(value).upper())
            values = row.get("symbols")
            if isinstance(values, list):
                row_codes.update(str(value).upper() for value in values)
            code_match = bool(row_codes and symbols and any(
                value in symbols or value.split(".")[0] in code_digits for value in row_codes
            ))
            if code_match:
                filtered.append(row)
                continue
            if resource in text_resources and (terms or code_digits):
                haystack = json.dumps(row, ensure_ascii=False, default=str)
                if not any(term in haystack for term in terms) and not any(code in haystack for code in code_digits):
                    continue
            elif row_codes and symbols:
                continue
            filtered.append(row)
        return filtered

    @staticmethod
    def _expand_zsxq_queries(values: List[Any]) -> List[str]:
        """Expand delimiter-separated planner phrases for literal full-text search."""
        expanded: List[str] = []
        for value in values:
            text = str(value or "").strip()
            if not text:
                continue
            expanded.append(text)
            expanded.extend(
                part.strip()
                for part in re.split(r"[\s,，、;；|/]+", text)
                if part.strip() and part.strip() != text
            )
        return list(dict.fromkeys(expanded))

    @staticmethod
    def _write_csv(path: Path, rows: List[Dict[str, Any]]) -> None:
        columns = list(dict.fromkeys(key for row in rows for key in row.keys()))
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=columns or ["message"])
            writer.writeheader()
            for row in rows:
                writer.writerow({key: DataAcquisitionService._cell(row.get(key)) for key in columns})

    @staticmethod
    def _write_sheet(workbook: Workbook, title: str, rows: List[Dict[str, Any]], error: Optional[str]) -> None:
        sheet = workbook.create_sheet(title=title or "dataset")
        if not rows:
            sheet.append(["status", "message"])
            sheet.append(["failed" if error else "empty", error or "No rows returned"])
            return
        columns = list(dict.fromkeys(key for row in rows for key in row.keys()))
        sheet.append(columns)
        for row in rows:
            sheet.append([DataAcquisitionService._cell(row.get(key)) for key in columns])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    @staticmethod
    def _cell(value: Any) -> Any:
        if isinstance(value, (dict, list, tuple)):
            return json.dumps(value, ensure_ascii=False, separators=(",", ":"))[:32000]
        if value is None:
            return ""
        return str(value)[:32000]

    @staticmethod
    def _safe_filename(value: str) -> str:
        cleaned = re.sub(r"[^\w\u4e00-\u9fff.-]+", "_", value, flags=re.UNICODE).strip("._")
        return cleaned[:60] or "dataset"

    @staticmethod
    def _safe_error(exc: Exception) -> str:
        if isinstance(exc, (DataAcquisitionError, FinancialDataValidationError)):
            return str(exc)[:300]
        return f"{type(exc).__name__}: 数据源调用失败"[:300]
