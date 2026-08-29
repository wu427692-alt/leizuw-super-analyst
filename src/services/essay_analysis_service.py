# -*- coding: utf-8 -*-
"""DeepSeek analysis, tagging and dashboard aggregation for research-note essays."""

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime, time as datetime_time, timedelta, timezone
import hashlib
import json
import logging
import os
from pathlib import Path
import re
import tempfile
import threading
import time
from typing import Any, Dict, Iterable, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
import requests
from json_repair import repair_json

from src.config import get_config
from src.data.stock_index_loader import get_stock_name_index_map
from src.repositories.essay_daily_report_repo import EssayDailyReportRepository
from src.repositories.essay_analysis_repo import EssayAnalysisRepository
from src.services.essay_market_insight_service import EssayMarketInsightService
from src.storage import utc_naive_now
from src.utils.essay_analysis_quality import is_low_quality_summary
from src.utils.essay_topic_taxonomy import TOPIC_TAXONOMY_VERSION, canonicalize_topics

logger = logging.getLogger(__name__)

ESSAY_PROMPT_VERSION = "essay-radar-v3-quality-retry"
DEFAULT_ESSAY_MODEL = "deepseek-v4-flash"
DAILY_REPORT_PROMPT_VERSION = "essay-daily-v4-longform-stock-candidates"

_CATEGORIES = {
    "company_research",
    "broker_view",
    "industry_chain",
    "macro_policy",
    "market_flow",
    "event_catalyst",
    "earnings",
    "risk_warning",
    "rumor",
    "other",
}
_SENTIMENTS = {"bullish", "bearish", "neutral", "mixed"}
_HORIZONS = {"intraday", "short", "medium", "long", "unclear"}
_STANCES = {"bullish", "bearish", "neutral", "watching"}
_TS_CODE_RE = re.compile(r"^(\d{6})(?:\.(SH|SS|SZ|BJ))?$", re.IGNORECASE)


def _bounded_env_int(name: str, default: int, *, minimum: int = 1, maximum: int = 256) -> int:
    try:
        value = int(os.getenv(name, str(default)))
    except (TypeError, ValueError):
        return default
    return max(minimum, min(value, maximum))


_DASHBOARD_ROWS_CACHE_TTL_SECONDS = 30.0
_DASHBOARD_ROWS_CACHE_MAX_ENTRIES = _bounded_env_int("ESSAY_DASHBOARD_ROWS_CACHE_MAX_ENTRIES", 4)
_dashboard_rows_cache: Dict[Any, tuple[float, List[Dict[str, Any]]]] = {}
_dashboard_rows_cache_lock = threading.Lock()
_DEEP_INSIGHTS_CACHE_TTL_SECONDS = 1800.0
_DEEP_INSIGHTS_CACHE_MAX_ENTRIES = _bounded_env_int("ESSAY_DEEP_INSIGHTS_CACHE_MAX_ENTRIES", 8)
_deep_insights_cache: Dict[tuple[Any, ...], tuple[float, Dict[str, Any]]] = {}
_deep_insights_cache_lock = threading.Lock()
_ESSAY_SUMMARY_CACHE_TTL_SECONDS = 30.0
_ESSAY_SUMMARY_CACHE_MAX_ENTRIES = _bounded_env_int("ESSAY_SUMMARY_CACHE_MAX_ENTRIES", 32)
_essay_summary_cache: Dict[tuple[Any, ...], tuple[float, Dict[str, Any]]] = {}
_essay_summary_cache_lock = threading.Lock()


def _prune_ttl_cache(cache: Dict[Any, tuple[float, Any]], *, now: float, ttl: float, max_entries: int) -> None:
    """Drop expired and oldest cache values while their owning lock is held."""
    expired = [key for key, (created_at, _) in cache.items() if now - created_at >= ttl]
    for key in expired:
        cache.pop(key, None)
    overflow = len(cache) - max_entries
    if overflow > 0:
        oldest = sorted(cache, key=lambda item: cache[item][0])[:overflow]
        for key in oldest:
            cache.pop(key, None)


def _cached_essay_summary(key: tuple[Any, ...], loader) -> Dict[str, Any]:
    now = time.monotonic()
    with _essay_summary_cache_lock:
        _prune_ttl_cache(
            _essay_summary_cache,
            now=now,
            ttl=_ESSAY_SUMMARY_CACHE_TTL_SECONDS,
            max_entries=_ESSAY_SUMMARY_CACHE_MAX_ENTRIES,
        )
        cached = _essay_summary_cache.get(key)
        if cached:
            return cached[1]
    value = loader()
    with _essay_summary_cache_lock:
        stored_at = time.monotonic()
        _essay_summary_cache[key] = (stored_at, value)
        _prune_ttl_cache(
            _essay_summary_cache,
            now=stored_at,
            ttl=_ESSAY_SUMMARY_CACHE_TTL_SECONDS,
            max_entries=_ESSAY_SUMMARY_CACHE_MAX_ENTRIES,
        )
    return value

_SYSTEM_PROMPT = """你是中国资本市场研究资料结构化分析员。请只根据输入文本提取事实、观点和风险，
不要补造公司、代码、数字或来源。每篇文本都必须输出一条结果，并严格输出一个 JSON object。

JSON 格式：
{
  "items": [
    {
      "topic_id": "原样返回",
      "summary": "不超过180字的中文摘要",
      "primary_category": "company_research|broker_view|industry_chain|macro_policy|market_flow|event_catalyst|earnings|risk_warning|rumor|other",
      "sentiment": "bullish|bearish|neutral|mixed",
      "time_horizon": "intraday|short|medium|long|unclear",
      "importance_score": 0,
      "confidence_score": 0.0,
      "tags": ["最多8个稳定中文标签"],
      "industries": ["最多5个申万风格行业名"],
      "themes": ["最多5个主题概念"],
      "stock_mentions": [
        {
          "ts_code": "明确时填写600519.SH；无法确认则为空字符串",
          "name": "公司或标的名称",
          "stance": "bullish|bearish|neutral|watching",
          "confidence": 0.0,
          "rationale": "不超过80字"
        }
      ],
      "key_points": ["最多5条"],
      "catalysts": ["最多4条"],
      "risks": ["最多4条"],
      "evidence": [{"claim":"核心判断", "evidence":"原文依据", "strength":"strong|medium|weak"}],
      "contradictions": ["原文内部矛盾、与常识冲突或尚未证实处，最多4条"],
      "falsification_conditions": ["什么事实出现会证伪该观点，最多4条"],
      "monitoring_points": ["后续应跟踪的数据、公告或时间节点，最多6条"],
      "earnings_impact": "对收入、利润、现金流或成本的可能传导；无依据写信息不足",
      "valuation_impact": "对估值逻辑和风险溢价的可能影响；无依据写信息不足",
      "source_quality": "high|medium|low|unknown",
      "novelty_score": 0,
      "information_type": "fact|management_guidance|institution_view|market_rumor|mixed|unknown"
    }
  ]
}

规则：importance_score 为0-100整数；confidence_score 和股票 confidence 为0-1。
summary 必须始终根据可见的标题和正文形成具体摘要，不得为空，不得写“信息不足，未形成有效摘要”或同义占位语。
即使正文很短，也要概括标题、可见事实和资料形态；只有图片或附件时，应明确写“原文主要为图片/附件，标题显示……”，而不是返回空摘要。
标签优先使用稳定词汇，例如“业绩超预期、涨价、产能扩张、国产替代、政策催化、机构观点、风险提示、订单增长、AI算力、机器人”。
必须把事实、公司指引、机构观点和市场传闻分开；结论要给原文证据、证伪条件与后续验证指标。
novelty_score 为0-100，衡量相对常见市场叙事的信息增量，不得把重复观点打高分。
券商观点不等于事实；传闻、转述、缺少来源或只有图片/文件标题时降低置信度并标记“信息不足”或“传闻”。
不得输出 Markdown、解释或 JSON 之外的文本。"""


class EssayAnalysisError(RuntimeError):
    """Safe error surfaced by the essay analysis pipeline."""


class DeepSeekEssayAnalyzer:
    """Batch analyzer bound directly to the official DeepSeek API."""

    def __init__(
        self,
        *,
        api_key: Optional[str] = None,
        base_url: Optional[str] = None,
        model: Optional[str] = None,
        timeout_seconds: Optional[int] = None,
        max_content_chars: Optional[int] = None,
        session: Optional[requests.Session] = None,
    ):
        config = get_config()
        keys = list(getattr(config, "deepseek_api_keys", None) or [])
        configured_key = str(api_key or (keys[0] if keys else getattr(config, "deepseek_api_key", "")) or "").strip()
        self.api_key = configured_key
        self.base_url = str(base_url or os.getenv("ESSAY_ANALYSIS_DEEPSEEK_BASE_URL") or "https://api.deepseek.com").rstrip("/")
        self.model = str(model or os.getenv("ESSAY_ANALYSIS_MODEL") or DEFAULT_ESSAY_MODEL).strip()
        self.timeout_seconds = max(
            10,
            min(int(timeout_seconds or os.getenv("ESSAY_ANALYSIS_TIMEOUT_SEC", "120")), 300),
        )
        self.max_content_chars = max(
            500,
            min(int(max_content_chars or os.getenv("ESSAY_ANALYSIS_MAX_CONTENT_CHARS", "5000")), 20000),
        )
        self.max_retries = max(1, min(int(os.getenv("ESSAY_ANALYSIS_MAX_RETRIES", "3")), 5))
        self.session = session or requests.Session()

    @property
    def configured(self) -> bool:
        return bool(self.api_key)

    def analyze_batch(self, notes: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
        if not self.configured:
            raise EssayAnalysisError("DEEPSEEK_API_KEY is not configured")
        if not notes:
            return {"items": [], "usage": {}, "raw_response": ""}

        payload_notes = [self._prompt_note(note) for note in notes]
        request_payload = {
            "model": self.model,
            "messages": [
                {"role": "system", "content": _SYSTEM_PROMPT},
                {
                    "role": "user",
                    "content": "请分析以下小作文数组并输出指定 JSON：\n" + json.dumps(payload_notes, ensure_ascii=False),
                },
            ],
            "response_format": {"type": "json_object"},
            "thinking": {"type": "disabled"},
            "temperature": 0.1,
            "max_tokens": 12000,
            "stream": False,
        }
        response_payload = self._post_with_retry(request_payload)
        raw_content = self._extract_content(response_payload)
        parsed = self._parse_json(raw_content)
        normalized_items = self._normalize_batch(parsed, notes)
        usage = response_payload.get("usage") if isinstance(response_payload.get("usage"), dict) else {}
        return {
            "items": normalized_items,
            "usage": {
                "prompt_tokens": int(usage.get("prompt_tokens") or 0),
                "completion_tokens": int(usage.get("completion_tokens") or 0),
                "total_tokens": int(usage.get("total_tokens") or 0),
            },
            "raw_response": raw_content,
        }

    def _post_with_retry(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        last_error = "DeepSeek request failed"
        for attempt in range(1, self.max_retries + 1):
            try:
                response = self.session.post(
                    f"{self.base_url}/chat/completions",
                    headers={
                        "Authorization": f"Bearer {self.api_key}",
                        "Content-Type": "application/json",
                    },
                    json=payload,
                    timeout=self.timeout_seconds,
                )
                if response.status_code == 200:
                    result = response.json()
                    if not isinstance(result, dict):
                        raise EssayAnalysisError("DeepSeek returned a non-object response")
                    return result
                retryable = response.status_code == 429 or response.status_code >= 500
                last_error = f"DeepSeek HTTP {response.status_code}"
                if not retryable:
                    raise EssayAnalysisError(last_error)
                retry_after = response.headers.get("Retry-After")
                delay = float(retry_after) if retry_after and retry_after.isdigit() else min(2 ** attempt, 10)
            except EssayAnalysisError:
                raise
            except (requests.RequestException, ValueError) as exc:
                last_error = f"DeepSeek request failed: {type(exc).__name__}"
                delay = min(2 ** attempt, 10)
            if attempt < self.max_retries:
                time.sleep(delay)
        raise EssayAnalysisError(last_error)

    @staticmethod
    def _extract_content(payload: Dict[str, Any]) -> str:
        try:
            content = payload["choices"][0]["message"]["content"]
        except (KeyError, IndexError, TypeError) as exc:
            raise EssayAnalysisError("DeepSeek response is missing message content") from exc
        normalized = str(content or "").strip()
        if not normalized:
            raise EssayAnalysisError("DeepSeek returned empty JSON content")
        return normalized

    @staticmethod
    def _parse_json(raw_content: str) -> Dict[str, Any]:
        cleaned = raw_content.strip()
        if cleaned.startswith("```"):
            cleaned = re.sub(r"^```(?:json)?\s*|\s*```$", "", cleaned, flags=re.IGNORECASE)
        try:
            parsed = json.loads(cleaned)
        except json.JSONDecodeError:
            try:
                parsed = json.loads(repair_json(cleaned))
            except (TypeError, ValueError, json.JSONDecodeError) as exc:
                raise EssayAnalysisError("DeepSeek returned invalid JSON") from exc
        if not isinstance(parsed, dict):
            raise EssayAnalysisError("DeepSeek JSON root must be an object")
        return parsed

    def _prompt_note(self, note: Dict[str, Any]) -> Dict[str, Any]:
        content = str(note.get("content") or "")
        return {
            "topic_id": str(note.get("topic_id") or ""),
            "title": str(note.get("title") or "")[:500],
            "content": content[: self.max_content_chars],
            "existing_symbols": list(note.get("existing_symbols") or []),
            "created_at": note.get("created_at"),
        }

    def _normalize_batch(
        self,
        payload: Dict[str, Any],
        notes: Sequence[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        raw_items = payload.get("items")
        if not isinstance(raw_items, list):
            raise EssayAnalysisError("DeepSeek JSON must contain an items array")
        note_by_topic = {str(note["topic_id"]): note for note in notes}
        normalized = []
        seen = set()
        for raw_item in raw_items:
            if not isinstance(raw_item, dict):
                continue
            topic_id = str(raw_item.get("topic_id") or "").strip()
            if topic_id not in note_by_topic or topic_id in seen:
                continue
            normalized_item = self._normalize_item(raw_item, note_by_topic[topic_id])
            if is_low_quality_summary(normalized_item["summary"]):
                logger.warning("[essay-radar] rejected low-quality summary for topic %s", topic_id)
                continue
            normalized.append(normalized_item)
            seen.add(topic_id)
        return normalized

    def _normalize_item(self, item: Dict[str, Any], note: Dict[str, Any]) -> Dict[str, Any]:
        category = str(item.get("primary_category") or "other").strip().lower()
        sentiment = str(item.get("sentiment") or "neutral").strip().lower()
        horizon = str(item.get("time_horizon") or "unclear").strip().lower()
        stock_mentions = []
        for raw_stock in item.get("stock_mentions") or []:
            if not isinstance(raw_stock, dict):
                continue
            name = str(raw_stock.get("name") or "").strip()[:100]
            ts_code = self._normalize_ts_code(raw_stock.get("ts_code"))
            if not name and not ts_code:
                continue
            stance = str(raw_stock.get("stance") or "neutral").strip().lower()
            stock_mentions.append({
                "ts_code": ts_code,
                "name": name,
                "stance": stance if stance in _STANCES else "neutral",
                "confidence": self._score(raw_stock.get("confidence"), maximum=1.0),
                "rationale": str(raw_stock.get("rationale") or "").strip()[:200],
            })
        existing_codes = {stock["ts_code"] for stock in stock_mentions if stock["ts_code"]}
        for existing_symbol in note.get("existing_symbols") or []:
            ts_code = self._normalize_ts_code(existing_symbol)
            if ts_code and ts_code not in existing_codes:
                stock_mentions.append({
                    "ts_code": ts_code,
                    "name": "",
                    "stance": "neutral",
                    "confidence": 0.5,
                    "rationale": "原文明确出现证券代码",
                })
        return {
            "topic_id": str(note["topic_id"]),
            "summary": str(item.get("summary") or "").strip()[:500],
            "primary_category": category if category in _CATEGORIES else "other",
            "sentiment": sentiment if sentiment in _SENTIMENTS else "neutral",
            "time_horizon": horizon if horizon in _HORIZONS else "unclear",
            "importance_score": int(round(self._score(item.get("importance_score"), maximum=100.0))),
            "confidence_score": self._score(item.get("confidence_score"), maximum=1.0),
            "tags": self._string_list(item.get("tags"), 8),
            "industries": self._string_list(item.get("industries"), 5),
            "themes": self._string_list(item.get("themes"), 5),
            "stock_mentions": stock_mentions[:10],
            "key_points": self._string_list(item.get("key_points"), 5, max_length=300),
            "catalysts": self._string_list(item.get("catalysts"), 4, max_length=300),
            "risks": self._string_list(item.get("risks"), 4, max_length=300),
            "evidence": self._evidence_list(item.get("evidence"), 6),
            "contradictions": self._string_list(item.get("contradictions"), 4, max_length=300),
            "falsification_conditions": self._string_list(item.get("falsification_conditions"), 4, max_length=300),
            "monitoring_points": self._string_list(item.get("monitoring_points"), 6, max_length=300),
            "earnings_impact": str(item.get("earnings_impact") or "信息不足").strip()[:500],
            "valuation_impact": str(item.get("valuation_impact") or "信息不足").strip()[:500],
            "source_quality": str(item.get("source_quality") or "unknown").strip().lower() if str(item.get("source_quality") or "unknown").strip().lower() in {"high", "medium", "low", "unknown"} else "unknown",
            "novelty_score": int(round(self._score(item.get("novelty_score"), maximum=100.0))),
            "information_type": str(item.get("information_type") or "unknown").strip().lower() if str(item.get("information_type") or "unknown").strip().lower() in {"fact", "management_guidance", "institution_view", "market_rumor", "mixed", "unknown"} else "unknown",
        }

    @staticmethod
    def _evidence_list(value: Any, limit: int) -> List[Dict[str, str]]:
        if not isinstance(value, list):
            return []
        result = []
        for raw in value:
            if not isinstance(raw, dict):
                continue
            claim = str(raw.get("claim") or "").strip()[:300]
            evidence = str(raw.get("evidence") or "").strip()[:500]
            strength = str(raw.get("strength") or "weak").strip().lower()
            if claim and evidence:
                result.append({"claim": claim, "evidence": evidence, "strength": strength if strength in {"strong", "medium", "weak"} else "weak"})
            if len(result) >= limit:
                break
        return result

    @staticmethod
    def _score(value: Any, *, maximum: float) -> float:
        try:
            number = float(value)
        except (TypeError, ValueError):
            return 0.0
        return round(max(0.0, min(number, maximum)), 4)

    @staticmethod
    def _string_list(value: Any, limit: int, *, max_length: int = 80) -> List[str]:
        if not isinstance(value, list):
            return []
        result = []
        seen = set()
        for item in value:
            normalized = str(item or "").strip()[:max_length]
            if not normalized or normalized in seen:
                continue
            seen.add(normalized)
            result.append(normalized)
            if len(result) >= limit:
                break
        return result

    @staticmethod
    def _normalize_ts_code(value: Any) -> str:
        raw = str(value or "").strip().upper()
        match = _TS_CODE_RE.fullmatch(raw)
        if not match:
            return ""
        digits, suffix = match.groups()
        suffix = (suffix or "").replace("SS", "SH")
        if not suffix:
            if digits.startswith(("4", "8")):
                suffix = "BJ"
            elif digits.startswith(("6", "9")):
                suffix = "SH"
            else:
                suffix = "SZ"
        return f"{digits}.{suffix}"


_DAILY_SYSTEM_PROMPT = """你是中国资本市场晨会主笔。仅依据给定的前一日机构段子结构化记录形成一份可审计的长篇日报。
必须区分事实、公司指引、机构观点和传闻，不得补造数据、股票、代码、价格或来源。报告应说明市场主线如何形成、
相互矛盾的证据、潜在盈利与估值传导、未来验证点。股票只能列为“重点研究候选”，不能写成确定收益或买卖指令；
每只候选必须来自输入中的明确股票提及，并引用 evidence_topic_ids。证据不足时宁可少列或不列。
严格输出一个 JSON object，不得输出 Markdown。所有长段落使用中文完整句子，可用两个换行分段。
格式：{
  "executive_summary":"800-1500字，至少4段：新增信息、主线结构、分歧风险、下一步观察",
  "market_regime":"100-250字市场叙事状态",
  "market_narrative":"600-1200字，解释主题之间的传导链和证据强弱",
  "key_themes":[{"name":"主题","count":1,"direction":"bullish|bearish|mixed|neutral","thesis":"150-300字逻辑","evidence":"事实与观点分层","counter_evidence":"反面证据或空缺","evidence_topic_ids":["id"]}],
  "stock_focus":[{"ts_code":"","name":"","mention_count":1,"stance":"bullish|bearish|mixed|neutral","conviction":"high|medium|low","time_horizon":"短期|中期|长期|不明确","thesis":"200-400字研究逻辑","why_now":"为何进入今日候选","earnings_path":"盈利传导与需要验证的数据","valuation_view":"估值影响与约束","catalysts":[],"risks":[],"validation_points":[],"evidence_topic_ids":[]}],
  "consensus":[], "divergences":[], "novel_signals":[], "earnings_implications":[], "valuation_implications":[],
  "risk_watch":[], "next_day_watchlist":[],
  "data_quality":{"coverage":"","limitations":[],"recommendation_rule":"候选仅来自明确提及且有证据链的股票"}
}"""


class EssayDailyReportService:
    """Generate and persist one previous-day synthesis per configured model."""

    def __init__(self, *, analysis_repo: Optional[EssayAnalysisRepository] = None, report_repo: Optional[EssayDailyReportRepository] = None):
        self.analysis_repo = analysis_repo or EssayAnalysisRepository()
        self.report_repo = report_repo or EssayDailyReportRepository()

    @staticmethod
    def configured_models() -> List[str]:
        raw = os.getenv("ESSAY_DAILY_REPORT_MODELS") or os.getenv("ESSAY_ANALYSIS_MODEL") or DEFAULT_ESSAY_MODEL
        return list(dict.fromkeys(value.strip() for value in raw.split(",") if value.strip()))[:20]

    def generate(self, *, report_date: Optional[str] = None, models: Optional[Sequence[str]] = None, force: bool = False) -> Dict[str, Any]:
        target = self._target_date(report_date)
        start, end = self._utc_bounds(target)
        rows = self.analysis_repo.completed_between(start=start, end=end)
        source_hash = hashlib.sha256(json.dumps(
            [(row["topic_id"], row.get("updated_at"), row.get("prompt_version")) for row in rows],
            ensure_ascii=False, separators=(",", ":"),
        ).encode("utf-8")).hexdigest()
        selected_models = list(models or self.configured_models())
        results = []
        for model in selected_models:
            existing = self.report_repo.get(target.isoformat(), model)
            if (existing and existing.get("status") == "completed" and
                    existing.get("source_hash") == source_hash and
                    existing.get("prompt_version") == DAILY_REPORT_PROMPT_VERSION and not force):
                results.append({"model": model, "status": "unchanged", "report": existing})
                continue
            self.report_repo.begin(report_date=target.isoformat(), model=model, prompt_version=DAILY_REPORT_PROMPT_VERSION, source_count=len(rows), source_hash=source_hash)
            if not rows:
                empty_report = self._empty_report(target)
                self.report_repo.save_success(report_date=target.isoformat(), model=model, report=empty_report, usage={})
                results.append({"model": model, "status": "completed", "source_count": 0})
                continue
            try:
                report, usage = self._call_model(model, target, rows)
                self.report_repo.save_success(report_date=target.isoformat(), model=model, report=report, usage=usage)
                results.append({"model": model, "status": "completed", "source_count": len(rows)})
            except Exception as exc:
                safe_error = f"{type(exc).__name__}: {str(exc)[:500]}"
                self.report_repo.save_failure(report_date=target.isoformat(), model=model, error=safe_error)
                results.append({"model": model, "status": "failed", "error": safe_error})
        return {"report_date": target.isoformat(), "source_count": len(rows), "models": results}

    def list(self, *, limit: int = 30, model: Optional[str] = None) -> Dict[str, Any]:
        items = self.report_repo.list(limit=limit, model=model)
        return {"items": items, "models": self.configured_models(), "total": len(items)}

    def _call_model(self, model: str, target: date, rows: Sequence[Dict[str, Any]]) -> tuple[Dict[str, Any], Dict[str, int]]:
        analyzer = DeepSeekEssayAnalyzer(model=model)
        if not analyzer.configured:
            raise EssayAnalysisError("DEEPSEEK_API_KEY is not configured")
        context = self._daily_context(rows)
        payload = {
            "model": model,
            "messages": [{"role": "system", "content": _DAILY_SYSTEM_PROMPT}, {"role": "user", "content": f"报告日期：{target.isoformat()}。全量统计覆盖全部 {len(rows)} 篇；代表性证据用于定性归纳。\n" + json.dumps(context, ensure_ascii=False)}],
            "response_format": {"type": "json_object"}, "thinking": {"type": "disabled"}, "temperature": 0.1,
            "max_tokens": 12000, "stream": False,
        }
        response = analyzer._post_with_retry(payload)
        report = analyzer._parse_json(analyzer._extract_content(response))
        report = self._normalize_report(report, rows)
        report["report_date"] = target.isoformat()
        report["source_count"] = len(rows)
        quality = report.get("data_quality") if isinstance(report.get("data_quality"), dict) else {}
        quality.update(context["coverage"])
        report["data_quality"] = quality
        usage = response.get("usage") if isinstance(response.get("usage"), dict) else {}
        return report, {key: int(usage.get(key) or 0) for key in ("prompt_tokens", "completion_tokens", "total_tokens")}

    @staticmethod
    def _normalize_report(report: Dict[str, Any], rows: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
        """Keep stock candidates tied to explicit source mentions and valid evidence ids."""
        valid_topic_ids = {str(row.get("topic_id") or "").strip() for row in rows}
        mentions_by_key: Dict[str, List[str]] = defaultdict(list)
        for row in rows:
            topic_id = str(row.get("topic_id") or "").strip()
            for mention in row.get("stock_mentions") or []:
                for value in (mention.get("ts_code"), mention.get("name")):
                    key = str(value or "").strip().casefold()
                    if key and topic_id and topic_id not in mentions_by_key[key]:
                        mentions_by_key[key].append(topic_id)

        normalized_stocks = []
        for candidate in report.get("stock_focus") or []:
            if not isinstance(candidate, dict):
                continue
            keys = {
                str(candidate.get("ts_code") or "").strip().casefold(),
                str(candidate.get("name") or "").strip().casefold(),
            } - {""}
            matched_topic_ids: List[str] = []
            for key in keys:
                matched_topic_ids.extend(mentions_by_key.get(key, []))
            matched_topic_ids = list(dict.fromkeys(matched_topic_ids))
            if not matched_topic_ids:
                continue
            cited = [
                str(value).strip() for value in (candidate.get("evidence_topic_ids") or [])
                if str(value).strip() in valid_topic_ids and str(value).strip() in matched_topic_ids
            ]
            candidate["evidence_topic_ids"] = list(dict.fromkeys(cited or matched_topic_ids[:8]))
            normalized_stocks.append(candidate)
        report["stock_focus"] = normalized_stocks[:12]

        for theme in report.get("key_themes") or []:
            if isinstance(theme, dict):
                theme["evidence_topic_ids"] = list(dict.fromkeys(
                    str(value).strip() for value in (theme.get("evidence_topic_ids") or [])
                    if str(value).strip() in valid_topic_ids
                ))[:12]
        return report

    @classmethod
    def _daily_context(cls, rows: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
        sentiment, categories, sources = Counter(), Counter(), Counter()
        information_types, source_quality, tags, themes, stocks = Counter(), Counter(), Counter(), Counter(), Counter()
        evidence_count = low_confidence = rumor_count = 0
        for row in rows:
            sentiment[row.get("sentiment") or "neutral"] += 1
            categories[row.get("primary_category") or "other"] += 1
            sources[(row.get("note") or {}).get("group_name") or "unknown"] += 1
            information_type = row.get("information_type") or "unknown"
            information_types[information_type] += 1
            source_quality[row.get("source_quality") or "unknown"] += 1
            tags.update(row.get("tags") or [])
            themes.update(row.get("themes") or [])
            evidence_count += int(bool(row.get("evidence")))
            low_confidence += int(float(row.get("confidence_score") or 0) < 0.5)
            rumor_count += int(information_type == "market_rumor" or row.get("primary_category") == "rumor")
            for mention in row.get("stock_mentions") or []:
                name = str(mention.get("name") or mention.get("ts_code") or "").strip()
                if name:
                    stocks[name] += 1
        ranked = sorted(rows, key=lambda row: (
            int(row.get("importance_score") or 0) + int(row.get("novelty_score") or 0),
            float(row.get("confidence_score") or 0),
        ), reverse=True)
        representative_rows = cls._representative_rows(ranked, limit=240)
        samples = []
        for row in representative_rows:
            samples.append({
                "topic_id": row["topic_id"], "title": (row.get("note") or {}).get("title"),
                "source": (row.get("note") or {}).get("group_name"), "summary": row.get("summary"),
                "category": row.get("primary_category"), "sentiment": row.get("sentiment"),
                "importance": row.get("importance_score"), "confidence": row.get("confidence_score"),
                "novelty": row.get("novelty_score"), "information_type": row.get("information_type"),
                "source_quality": row.get("source_quality"), "tags": row.get("tags"),
                "themes": row.get("themes"), "stocks": row.get("stock_mentions"),
                "key_points": row.get("key_points"), "catalysts": row.get("catalysts"),
                "risks": row.get("risks"), "evidence": row.get("evidence"),
                "contradictions": row.get("contradictions"), "falsification_conditions": row.get("falsification_conditions"),
                "monitoring_points": row.get("monitoring_points"),
            })
        representative_sources = {
            str((row.get("note") or {}).get("group_name") or "unknown") for row in representative_rows
        }
        representative_categories = {
            str(row.get("primary_category") or "other") for row in representative_rows
        }
        watchlist_records = sum(1 for row in representative_rows if cls._matches_daily_watchlist(row))
        return {
            "coverage": {
                "total_records": len(rows), "representative_records": len(samples),
                "evidence_records": evidence_count,
                "evidence_coverage_percent": round(evidence_count / len(rows) * 100, 1) if rows else 0,
                "low_confidence_records": low_confidence, "rumor_records": rumor_count,
                "representative_watchlist_records": watchlist_records,
                "representative_sources": len(representative_sources),
                "representative_categories": len(representative_categories),
                "selection_strategy": "watchlist_priority_plus_source_category_information_type_strata",
            },
            "full_population_aggregates": {
                "sentiment": dict(sentiment), "categories": dict(categories), "sources": dict(sources),
                "information_types": dict(information_types), "source_quality": dict(source_quality),
                "top_tags": tags.most_common(30), "top_themes": themes.most_common(30),
                "top_stocks": stocks.most_common(50),
            },
            "representative_evidence": samples,
        }

    @classmethod
    def _representative_rows(cls, ranked: Sequence[Dict[str, Any]], *, limit: int) -> List[Dict[str, Any]]:
        """Keep watchlist evidence and diverse strata before filling by analytical value."""
        safe_limit = max(1, min(int(limit), 500))
        selected: Dict[str, Dict[str, Any]] = {}

        def add(row: Dict[str, Any]) -> None:
            if len(selected) >= safe_limit:
                return
            topic_id = str(row.get("topic_id") or "").strip()
            if topic_id:
                selected.setdefault(topic_id, row)

        watchlist = cls._daily_watchlist()
        per_stock_limit = max(8, min(60, 120 // max(1, len(watchlist))))
        for symbol, name in watchlist:
            matched = [row for row in ranked if cls._row_mentions(row, symbol, name)]
            for row in matched[:per_stock_limit]:
                add(row)

        strata: Dict[tuple[str, str], int] = defaultdict(int)
        for row in ranked:
            values = (
                ("source", str((row.get("note") or {}).get("group_name") or "unknown")),
                ("category", str(row.get("primary_category") or "other")),
                ("information_type", str(row.get("information_type") or "unknown")),
            )
            for dimension, value in values:
                key = (dimension, value)
                if strata[key] < 4:
                    add(row)
                    strata[key] += 1
        for row in ranked:
            add(row)
        return list(selected.values())

    @staticmethod
    def _daily_watchlist() -> List[tuple[str, str]]:
        raw = os.getenv("ESSAY_WATCHLIST") or "603306.SH:华懋科技,300476.SZ:胜宏科技"
        result = []
        for item in raw.split(","):
            symbol, _, name = item.strip().partition(":")
            if symbol and name:
                result.append((symbol.strip().upper(), name.strip()))
        return result[:20]

    @staticmethod
    def _row_mentions(row: Dict[str, Any], symbol: str, name: str) -> bool:
        return any(
            symbol.upper() == str(item.get("ts_code") or "").upper()
            or name.lower() in str(item.get("name") or "").lower()
            for item in (row.get("stock_mentions") or [])
        )

    @classmethod
    def _matches_daily_watchlist(cls, row: Dict[str, Any]) -> bool:
        return any(cls._row_mentions(row, symbol, name) for symbol, name in cls._daily_watchlist())

    @staticmethod
    def _target_date(value: Optional[str]) -> date:
        if value:
            return date.fromisoformat(value)
        return (datetime.now(timezone(timedelta(hours=8))) - timedelta(days=1)).date()

    @staticmethod
    def _utc_bounds(target: date) -> tuple[datetime, datetime]:
        shanghai = timezone(timedelta(hours=8))
        local_start = datetime.combine(target, datetime_time.min, tzinfo=shanghai)
        local_end = local_start + timedelta(days=1)
        return local_start.astimezone(timezone.utc).replace(tzinfo=None), local_end.astimezone(timezone.utc).replace(tzinfo=None)

    @staticmethod
    def _empty_report(target: date) -> Dict[str, Any]:
        return {"report_date": target.isoformat(), "source_count": 0, "executive_summary": "前一日没有已完成分析的新增小作文。", "market_regime": "无新增样本", "key_themes": [], "stock_focus": [], "consensus": [], "divergences": [], "novel_signals": [], "earnings_implications": [], "valuation_implications": [], "risk_watch": [], "next_day_watchlist": [], "data_quality": {"coverage": "无新增样本", "limitations": ["没有可汇总记录"]}}


class EssayAnalysisService:
    """Application service for queue control, query and aggregate stock views."""

    def __init__(self, repository: Optional[EssayAnalysisRepository] = None):
        self.repo = repository or EssayAnalysisRepository()
        self.report_repo = EssayDailyReportRepository(self.repo.db)
        self.model = str(os.getenv("ESSAY_ANALYSIS_MODEL") or DEFAULT_ESSAY_MODEL).strip()

    def enqueue_recent(self, *, days: int = 30) -> Dict[str, Any]:
        safe_days = max(1, min(int(days), 3650))
        cutoff = utc_naive_now() - timedelta(days=safe_days)
        result = self.repo.enqueue_recent(
            cutoff=cutoff,
            model=self.model,
            prompt_version=ESSAY_PROMPT_VERSION,
        )
        return {"days": safe_days, "cutoff": cutoff.isoformat() + "Z", **result}

    def enqueue_topic_ids(self, topic_ids: Iterable[str]) -> Dict[str, int]:
        return self.repo.enqueue_topic_ids(
            topic_ids,
            model=self.model,
            prompt_version=ESSAY_PROMPT_VERSION,
        )

    def enqueue_unqueued(self, *, count: int, order: str = "newest") -> Dict[str, Any]:
        return self.repo.enqueue_unqueued(
            limit=max(1, min(int(count), 5000)),
            order=order,
            model=self.model,
            prompt_version=ESSAY_PROMPT_VERSION,
        )

    def historical_backlog(self) -> Dict[str, Any]:
        if type(self.repo) is not EssayAnalysisRepository:
            return self.repo.historical_backlog()
        database_key = str(getattr(self.repo.db, "_db_url", id(self.repo.db)))
        return _cached_essay_summary(
            ("historical-backlog", database_key, self.repo.cache_revision()),
            self.repo.historical_backlog,
        )

    def list_analyses(self, **filters: Any) -> Dict[str, Any]:
        days = max(1, min(int(filters.pop("days", 30) or 30), 3650))
        page = max(1, int(filters.get("page") or 1))
        page_size = max(1, min(int(filters.get("page_size") or 20), 100))
        cutoff = utc_naive_now() - timedelta(days=days)
        rows, total = self.repo.list_analyses(cutoff=cutoff, **filters)
        return {"items": rows, "total": total, "page": page, "page_size": page_size}

    def list_feed(self, **filters: Any) -> Dict[str, Any]:
        days = max(0, min(int(filters.pop("days", 0) or 0), 3650))
        page = max(1, int(filters.get("page") or 1))
        page_size = max(1, min(int(filters.get("page_size") or 20), 100))
        cutoff = utc_naive_now() - timedelta(days=days) if days else None
        rows, total = self.repo.list_feed(cutoff=cutoff, **filters)
        return {
            "items": rows,
            "total": total,
            "page": page,
            "page_size": page_size,
            "scope": "all_stored_notes" if cutoff is None else f"last_{days}_days",
        }

    def export_feed_excel(self, **filters: Any) -> Dict[str, Any]:
        """Export every row matching the information-feed filters to one XLSX."""
        days = max(0, min(int(filters.pop("days", 0) or 0), 3650))
        selected_topic_ids = list(dict.fromkeys(
            str(value).strip() for value in (filters.pop("topic_ids", None) or []) if str(value).strip()
        ))
        cutoff = utc_naive_now() - timedelta(days=days) if days else None
        export_filters = {
            key: filters.get(key)
            for key in (
                "query", "query_scope", "analysis_status", "sentiment", "category", "tag", "stock", "min_importance",
            )
        }
        workbook = Workbook(write_only=True)
        result_sheet = workbook.create_sheet("搜索结果与分析标签")
        raw_sheet = workbook.create_sheet("原文全文")
        condition_sheet = workbook.create_sheet("导出条件")

        result_headers = [
            "序号", "小作文ID", "发布时间", "知识星球", "作者", "标题", "原文摘要",
            "AI状态", "模型", "AI摘要", "类型", "情绪", "时间周期", "重要度", "置信度",
            "标签", "行业", "主题", "股票提及", "核心要点", "催化", "风险", "证据",
            "矛盾点", "证伪条件", "跟踪点", "盈利影响", "估值影响", "来源质量", "新颖度",
            "信息类型", "附件", "图片",
        ]
        raw_headers = ["小作文ID", "标题", "发布时间", "分段", "总分段", "原文"]
        self._append_excel_header(result_sheet, result_headers)
        self._append_excel_header(raw_sheet, raw_headers)
        result_sheet.freeze_panes = "A2"
        raw_sheet.freeze_panes = "A2"

        row_count = 0
        raw_chunk_count = 0
        for row_count, item in enumerate(
            self.repo.iter_feed(
                cutoff=cutoff,
                topic_ids=selected_topic_ids if selected_topic_ids else None,
                **export_filters,
            ),
            start=1,
        ):
            note = item.get("note") or {}
            content = str(note.get("content") or "")
            chunks = self._excel_chunks(content)
            result_sheet.append([
                row_count,
                self._excel_cell(item.get("topic_id")),
                self._excel_cell(note.get("created_at")),
                self._excel_cell(note.get("group_name")),
                self._excel_cell(note.get("author_name")),
                self._excel_cell(note.get("title")),
                self._excel_cell(content[:500]),
                self._excel_cell(item.get("status")),
                self._excel_cell(item.get("model")),
                self._excel_cell(item.get("summary")),
                self._excel_cell(item.get("primary_category")),
                self._excel_cell(item.get("sentiment")),
                self._excel_cell(item.get("time_horizon")),
                item.get("importance_score"),
                item.get("confidence_score"),
                self._excel_cell(item.get("tags")),
                self._excel_cell(item.get("industries")),
                self._excel_cell(item.get("themes")),
                self._excel_cell(item.get("stock_mentions")),
                self._excel_cell(item.get("key_points")),
                self._excel_cell(item.get("catalysts")),
                self._excel_cell(item.get("risks")),
                self._excel_cell(item.get("evidence")),
                self._excel_cell(item.get("contradictions")),
                self._excel_cell(item.get("falsification_conditions")),
                self._excel_cell(item.get("monitoring_points")),
                self._excel_cell(item.get("earnings_impact")),
                self._excel_cell(item.get("valuation_impact")),
                self._excel_cell(item.get("source_quality")),
                item.get("novelty_score"),
                self._excel_cell(item.get("information_type")),
                self._excel_cell(note.get("files")),
                self._excel_cell(note.get("images")),
            ])
            for chunk_index, chunk in enumerate(chunks, start=1):
                raw_sheet.append([
                    self._excel_cell(item.get("topic_id")),
                    self._excel_cell(note.get("title")),
                    self._excel_cell(note.get("created_at")),
                    chunk_index,
                    len(chunks),
                    self._excel_cell(chunk),
                ])
                raw_chunk_count += 1

        condition_rows = [
            ("导出时间", datetime.now().astimezone().isoformat(timespec="seconds")),
            ("匹配数量", row_count),
            ("原文分段数量", raw_chunk_count),
            ("时间范围", f"近 {days} 日" if days else "全部已入库"),
            ("关键词", export_filters.get("query") or ""),
            ("检索范围", "仅标题" if export_filters.get("query_scope") == "title" else "全文与分析标签"),
            ("AI状态", export_filters.get("analysis_status") or "全部"),
            ("情绪", export_filters.get("sentiment") or "全部"),
            ("类型", export_filters.get("category") or "全部"),
            ("标签", export_filters.get("tag") or "全部"),
            ("股票", export_filters.get("stock") or "全部"),
            ("最低重要度", export_filters.get("min_importance") or 0),
            ("勾选小作文", len(selected_topic_ids) if selected_topic_ids else "按当前筛选条件全量导出"),
            ("说明", "搜索结果工作表含分析标签和原文摘要；原文全文工作表按 Excel 单元格上限分段保存完整正文。"),
        ]
        self._append_excel_header(condition_sheet, ["项目", "值"])
        for key, value in condition_rows:
            condition_sheet.append([self._excel_cell(key), self._excel_cell(value)])
        condition_sheet.freeze_panes = "A2"

        handle = tempfile.NamedTemporaryFile(prefix="essay-feed-export-", suffix=".xlsx", delete=False)
        handle.close()
        path = Path(handle.name)
        try:
            workbook.save(path)
        except Exception:
            path.unlink(missing_ok=True)
            raise
        suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
        return {
            "path": str(path),
            "filename": f"{'小作文已选原文' if selected_topic_ids else '小作文检索结果'}_{suffix}.xlsx",
            "row_count": row_count,
            "raw_chunk_count": raw_chunk_count,
        }

    @staticmethod
    def _append_excel_header(sheet: Any, labels: Sequence[str]) -> None:
        cells = []
        for label in labels:
            cell = WriteOnlyCell(sheet, value=label)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="155E75")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cells.append(cell)
        sheet.append(cells)

    @staticmethod
    def _excel_chunks(value: str, size: int = 30000) -> List[str]:
        cleaned = ILLEGAL_CHARACTERS_RE.sub("", str(value or ""))
        return [cleaned[index:index + size] for index in range(0, len(cleaned), size)] or [""]

    @staticmethod
    def _excel_cell(value: Any) -> Any:
        if value is None:
            return ""
        if isinstance(value, (dict, list, tuple, set)):
            text = json.dumps(value, ensure_ascii=False, default=str)
        else:
            text = str(value)
        text = ILLEGAL_CHARACTERS_RE.sub("", text)
        if text.startswith(("=", "+", "-", "@", "\t", "\r")):
            text = "'" + text
        return text[:32760]

    def get_analysis(self, topic_id: str) -> Dict[str, Any]:
        row = self.repo.get_analysis(str(topic_id or "").strip())
        if row is None:
            raise KeyError(f"essay analysis not found: {topic_id}")
        return row

    def progress(self, *, days: int = 30) -> Dict[str, Any]:
        safe_days = max(1, min(int(days), 3650))
        cutoff = utc_naive_now() - timedelta(days=safe_days)
        analyzer = DeepSeekEssayAnalyzer(model=self.model)
        database_key = str(getattr(self.repo.db, "_db_url", id(self.repo.db)))
        progress = self.repo.progress(cutoff=cutoff) if type(self.repo) is not EssayAnalysisRepository else _cached_essay_summary(
            ("progress", database_key, safe_days, self.repo.cache_revision()),
            lambda: self.repo.progress(cutoff=cutoff),
        )
        return {
            "days": safe_days,
            "model": self.model,
            "deepseek_configured": analyzer.configured,
            **progress,
        }

    def _completed_dashboard_rows(self, safe_days: int) -> List[Dict[str, Any]]:
        """Share the expensive decoded row snapshot across dashboard consumers."""
        cutoff = utc_naive_now() - timedelta(days=safe_days)
        # Test doubles must remain isolated and deterministic.
        if type(self.repo) is not EssayAnalysisRepository:
            return self.repo.completed_for_dashboard(cutoff=cutoff)
        now = time.monotonic()
        with _dashboard_rows_cache_lock:
            _prune_ttl_cache(
                _dashboard_rows_cache,
                now=now,
                ttl=_DASHBOARD_ROWS_CACHE_TTL_SECONDS,
                max_entries=_DASHBOARD_ROWS_CACHE_MAX_ENTRIES,
            )
            cached = _dashboard_rows_cache.get(safe_days)
            if cached:
                return cached[1]
        rows = self.repo.completed_for_dashboard(cutoff=cutoff)
        with _dashboard_rows_cache_lock:
            stored_at = time.monotonic()
            _dashboard_rows_cache[safe_days] = (stored_at, rows)
            _prune_ttl_cache(
                _dashboard_rows_cache,
                now=stored_at,
                ttl=_DASHBOARD_ROWS_CACHE_TTL_SECONDS,
                max_entries=_DASHBOARD_ROWS_CACHE_MAX_ENTRIES,
            )
        return rows

    def _completed_dashboard_rows_between(self, start_day: date, end_day: date) -> List[Dict[str, Any]]:
        """Read one exact Shanghai-calendar window and share the decoded snapshot."""
        shanghai = timezone(timedelta(hours=8))
        start = datetime.combine(start_day, datetime_time.min, tzinfo=shanghai).astimezone(timezone.utc).replace(tzinfo=None)
        end = datetime.combine(end_day + timedelta(days=1), datetime_time.min, tzinfo=shanghai).astimezone(timezone.utc).replace(tzinfo=None)
        if type(self.repo) is not EssayAnalysisRepository:
            return self.repo.completed_between(start=start, end=end)
        cache_key = ("range", start_day.isoformat(), end_day.isoformat())
        now = time.monotonic()
        with _dashboard_rows_cache_lock:
            _prune_ttl_cache(
                _dashboard_rows_cache,
                now=now,
                ttl=_DASHBOARD_ROWS_CACHE_TTL_SECONDS,
                max_entries=_DASHBOARD_ROWS_CACHE_MAX_ENTRIES,
            )
            cached = _dashboard_rows_cache.get(cache_key)
            if cached:
                return cached[1]
        rows = self.repo.completed_between(start=start, end=end)
        with _dashboard_rows_cache_lock:
            stored_at = time.monotonic()
            _dashboard_rows_cache[cache_key] = (stored_at, rows)
            _prune_ttl_cache(
                _dashboard_rows_cache,
                now=stored_at,
                ttl=_DASHBOARD_ROWS_CACHE_TTL_SECONDS,
                max_entries=_DASHBOARD_ROWS_CACHE_MAX_ENTRIES,
            )
        return rows

    def dashboard(self, *, days: int = 30, top_n: int = 12) -> Dict[str, Any]:
        safe_days = max(1, min(int(days), 3650))
        safe_top_n = max(3, min(int(top_n), 30))
        if type(self.repo) is EssayAnalysisRepository:
            database_key = str(getattr(self.repo.db, "_db_url", id(self.repo.db)))
            cache_key = (
                "dashboard", database_key, safe_days, safe_top_n,
                self.repo.cache_revision(),
            )
            with _essay_summary_cache_lock:
                now = time.monotonic()
                _prune_ttl_cache(
                    _essay_summary_cache,
                    now=now,
                    ttl=_ESSAY_SUMMARY_CACHE_TTL_SECONDS,
                    max_entries=_ESSAY_SUMMARY_CACHE_MAX_ENTRIES,
                )
                cached = _essay_summary_cache.get(cache_key)
                if cached:
                    return cached[1]
        rows = self._completed_dashboard_rows(safe_days)
        sentiment = Counter()
        categories = Counter()
        horizons = Counter()
        tags = Counter()
        industries = Counter()
        themes = Counter()
        tag_by_day: Dict[str, Counter] = defaultdict(Counter)
        stocks: Dict[str, Dict[str, Any]] = {}
        importance_total = 0
        token_total = 0
        for row in rows:
            sentiment[row.get("sentiment") or "neutral"] += 1
            categories[row.get("primary_category") or "other"] += 1
            horizons[row.get("time_horizon") or "unclear"] += 1
            importance = int(row.get("importance_score") or 0)
            importance_total += importance
            token_total += int(row.get("total_tokens") or 0)
            note_date = str((row.get("note") or {}).get("created_at") or "")[:10]
            for tag in row.get("tags") or []:
                tags[tag] += 1
                if note_date:
                    tag_by_day[note_date][tag] += 1
            industries.update(row.get("industries") or [])
            themes.update(row.get("themes") or [])
            for mention in row.get("stock_mentions") or []:
                key = str(mention.get("ts_code") or mention.get("name") or "").strip()
                if not key:
                    continue
                bucket = stocks.setdefault(key, {
                    "key": key,
                    "ts_code": mention.get("ts_code") or "",
                    "name": mention.get("name") or "",
                    "mention_count": 0,
                    "bullish": 0,
                    "bearish": 0,
                    "neutral": 0,
                    "watching": 0,
                    "importance_total": 0,
                    "confidence_total": 0.0,
                    "latest_at": None,
                })
                bucket["mention_count"] += 1
                stance = mention.get("stance") if mention.get("stance") in _STANCES else "neutral"
                bucket[stance] += 1
                bucket["importance_total"] += importance
                bucket["confidence_total"] += float(mention.get("confidence") or 0)
                created_at = (row.get("note") or {}).get("created_at")
                if created_at and (not bucket["latest_at"] or created_at > bucket["latest_at"]):
                    bucket["latest_at"] = created_at

        top_tags = [name for name, _ in tags.most_common(safe_top_n)]
        stock_rows = []
        for bucket in stocks.values():
            count = bucket["mention_count"]
            bucket["average_importance"] = round(bucket.pop("importance_total") / count, 1)
            bucket["average_confidence"] = round(bucket.pop("confidence_total") / count, 3)
            stock_rows.append(bucket)
        stock_rows.sort(
            key=lambda item: (item["mention_count"], item["average_importance"]),
            reverse=True,
        )
        highlights = sorted(
            rows,
            key=lambda item: (item.get("importance_score") or 0, (item.get("note") or {}).get("created_at") or ""),
            reverse=True,
        )[:8]
        payload = {
            "days": safe_days,
            "generated_at": utc_naive_now().isoformat() + "Z",
            "summary": {
                "analyzed_count": len(rows),
                "average_importance": round(importance_total / len(rows), 1) if rows else 0.0,
                "stock_count": len(stocks),
                "tag_count": len(tags),
                "total_tokens": token_total,
            },
            "sentiment": self._counter_rows(sentiment),
            "categories": self._counter_rows(categories),
            "horizons": self._counter_rows(horizons),
            "top_tags": self._counter_rows(tags, safe_top_n),
            "top_industries": self._counter_rows(industries, safe_top_n),
            "top_themes": self._counter_rows(themes, safe_top_n),
            "top_stocks": stock_rows[:safe_top_n],
            "tag_trend": [
                {"date": day, "tag": tag, "count": count}
                for day in sorted(tag_by_day)
                for tag, count in tag_by_day[day].items()
                if tag in top_tags
            ],
            "highlights": highlights,
        }
        if type(self.repo) is EssayAnalysisRepository:
            with _essay_summary_cache_lock:
                stored_at = time.monotonic()
                _essay_summary_cache[cache_key] = (stored_at, payload)
                _prune_ttl_cache(
                    _essay_summary_cache,
                    now=stored_at,
                    ttl=_ESSAY_SUMMARY_CACHE_TTL_SECONDS,
                    max_entries=_ESSAY_SUMMARY_CACHE_MAX_ENTRIES,
                )
            return payload

    def insights(self, *, days: int = 30, trend_days: int = 14) -> Dict[str, Any]:
        """Evidence, trend, model and watchlist cockpit built from completed analyses."""
        safe_days = max(1, min(int(days), 3650))
        safe_trend_days = max(7, min(int(trend_days), 90))
        rows = self._completed_dashboard_rows(safe_days)
        shanghai = timezone(timedelta(hours=8))
        today = datetime.now(shanghai).date()
        yesterday = today - timedelta(days=1)
        trend_start = today - timedelta(days=safe_trend_days - 1)

        def local_day(row: Dict[str, Any]) -> Optional[date]:
            raw = str((row.get("note") or {}).get("created_at") or "")
            if not raw:
                return None
            try:
                parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
                if parsed.tzinfo is None:
                    parsed = parsed.replace(tzinfo=timezone.utc)
                return parsed.astimezone(shanghai).date()
            except ValueError:
                return None

        quality, information_types, sources = Counter(), Counter(), Counter()
        daily: Dict[date, Dict[str, Any]] = {}
        yesterday_rows: List[Dict[str, Any]] = []
        evidence_records = 0
        for row in rows:
            row_day = local_day(row)
            if row_day == yesterday:
                yesterday_rows.append(row)
            quality[row.get("source_quality") or "unknown"] += 1
            information_types[row.get("information_type") or "unknown"] += 1
            sources[(row.get("note") or {}).get("group_name") or "unknown"] += 1
            evidence_records += int(bool(row.get("evidence")))
            if row_day and row_day >= trend_start:
                bucket = daily.setdefault(row_day, {
                    "date": row_day.isoformat(), "total": 0, "bullish": 0, "bearish": 0,
                    "neutral": 0, "mixed": 0, "importance_total": 0, "confidence_total": 0.0,
                })
                bucket["total"] += 1
                sentiment = row.get("sentiment") if row.get("sentiment") in _SENTIMENTS else "neutral"
                bucket[sentiment] += 1
                bucket["importance_total"] += int(row.get("importance_score") or 0)
                bucket["confidence_total"] += float(row.get("confidence_score") or 0)

        trend = []
        for offset in range(safe_trend_days):
            day = trend_start + timedelta(days=offset)
            bucket = daily.get(day, {"date": day.isoformat(), "total": 0, "bullish": 0, "bearish": 0, "neutral": 0, "mixed": 0, "importance_total": 0, "confidence_total": 0.0})
            total = bucket["total"]
            trend.append({
                **{key: bucket[key] for key in ("date", "total", "bullish", "bearish", "neutral", "mixed")},
                "average_importance": round(bucket["importance_total"] / total, 1) if total else 0,
                "average_confidence": round(bucket["confidence_total"] / total, 3) if total else 0,
            })

        latest_reports = self.report_repo.list(limit=60)
        latest_date = max((item["report_date"] for item in latest_reports), default=None)
        report_rows = [item for item in latest_reports if item["report_date"] == latest_date] if latest_date else []
        completed_report_rows = [item for item in report_rows if item.get("status") == "completed"]
        comparison = self._compare_reports(completed_report_rows)

        watchlist = []
        for symbol, name in self._configured_watchlist():
            matched = []
            for row in rows:
                mentions = row.get("stock_mentions") or []
                if any(symbol.upper() == str(item.get("ts_code") or "").upper() or
                       name.lower() in str(item.get("name") or "").lower() for item in mentions):
                    matched.append(row)
            watchlist.append(self._watchlist_summary(symbol, name, matched, local_day, today, safe_trend_days))

        low_confidence = sum(1 for row in yesterday_rows if float(row.get("confidence_score") or 0) < 0.5)
        rumors = sum(1 for row in yesterday_rows if row.get("information_type") == "market_rumor" or row.get("primary_category") == "rumor")
        high_importance = sum(1 for row in yesterday_rows if int(row.get("importance_score") or 0) >= 80)
        high_novelty = sorted(yesterday_rows, key=lambda row: (int(row.get("novelty_score") or 0), int(row.get("importance_score") or 0)), reverse=True)[:8]
        return {
            "generated_at": utc_naive_now().isoformat() + "Z",
            "window_days": safe_days,
            "latest_data_at": max((str((row.get("note") or {}).get("created_at") or "") for row in rows), default=None),
            "yesterday": {
                "date": yesterday.isoformat(), "analyzed_count": len(yesterday_rows),
                "high_importance_count": high_importance, "low_confidence_count": low_confidence,
                "rumor_count": rumors,
                "evidence_coverage_percent": round(sum(bool(row.get("evidence")) for row in yesterday_rows) / len(yesterday_rows) * 100, 1) if yesterday_rows else 0,
            },
            "coverage": {
                "analyzed_count": len(rows), "evidence_records": evidence_records,
                "evidence_coverage_percent": round(evidence_records / len(rows) * 100, 1) if rows else 0,
                "configured_models": EssayDailyReportService.configured_models(),
                "completed_report_models": len(completed_report_rows),
            },
            "trend": trend,
            "source_quality": self._counter_rows(quality),
            "information_types": self._counter_rows(information_types),
            "source_mix": self._counter_rows(sources, 20),
            "model_comparison": {
                "report_date": latest_date, "reports": report_rows,
                "consensus": comparison["consensus"],
                "divergences": comparison["divergences"],
            },
            "watchlist": watchlist,
            "high_novelty_signals": high_novelty,
        }

    def deep_insights(
        self,
        *,
        days: int = 30,
        trend_days: int = 14,
        horizon: Optional[str] = None,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Build a date-bounded source → theme → stock → market-evidence atlas."""
        shanghai = timezone(timedelta(hours=8))
        today = datetime.now(shanghai).date()
        preset_days = {"short": 14, "medium": 90, "long": 180}
        normalized_horizon = str(horizon or "").strip().lower()
        try:
            selected_end = date.fromisoformat(end_date) if end_date else today
            if normalized_horizon == "custom":
                if not start_date or not end_date:
                    raise ValueError("自定义窗口必须同时提供开始和结束日期")
                selected_start = date.fromisoformat(start_date)
            elif normalized_horizon in preset_days:
                selected_start = selected_end - timedelta(days=preset_days[normalized_horizon] - 1)
            else:
                safe_days = max(7, min(int(days), 3650))
                normalized_horizon = "custom"
                selected_start = selected_end - timedelta(days=safe_days - 1)
        except ValueError as exc:
            raise ValueError("日期必须使用 YYYY-MM-DD，且自定义窗口需要完整起止日期") from exc
        if selected_start > selected_end:
            raise ValueError("开始日期不能晚于结束日期")
        if (selected_end - selected_start).days >= 3650:
            raise ValueError("洞察窗口最长为 3650 天")
        safe_days = (selected_end - selected_start).days + 1
        safe_trend_days = max(7, min(int(trend_days), 90))
        if safe_days <= 31:
            granularity = "day"
        elif safe_days <= 180:
            granularity = "week"
        else:
            granularity = "month"
        cache_key = (
            id(self.repo.db), normalized_horizon, selected_start.isoformat(),
            selected_end.isoformat(), granularity,
        )
        now = time.monotonic()
        with _deep_insights_cache_lock:
            _prune_ttl_cache(
                _deep_insights_cache,
                now=now,
                ttl=_DEEP_INSIGHTS_CACHE_TTL_SECONDS,
                max_entries=_DEEP_INSIGHTS_CACHE_MAX_ENTRIES,
            )
            cached = _deep_insights_cache.get(cache_key)
            if cached:
                return cached[1]
        rows = self._completed_dashboard_rows_between(selected_start, selected_end)
        trend_start = max(selected_start, selected_end - timedelta(days=safe_trend_days - 1))
        comparison_days = 7 if safe_days <= 31 else 30 if safe_days <= 180 else 90
        current_start = max(selected_start, selected_end - timedelta(days=comparison_days - 1))
        previous_start = max(selected_start, current_start - timedelta(days=comparison_days))

        def local_day(row: Dict[str, Any]) -> Optional[date]:
            raw = str((row.get("note") or {}).get("created_at") or "")
            if not raw:
                return None
            try:
                parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
                if parsed.tzinfo is None:
                    parsed = parsed.replace(tzinfo=timezone.utc)
                return parsed.astimezone(shanghai).date()
            except ValueError:
                return None

        sources: Counter[str] = Counter()
        themes: Counter[str] = Counter()
        stocks: Counter[str] = Counter()
        signals: Counter[tuple[str, str]] = Counter()
        source_theme: Counter[tuple[str, str]] = Counter()
        theme_stock: Counter[tuple[str, str]] = Counter()
        stock_signal: Counter[tuple[str, str, str]] = Counter()
        theme_by_day: Dict[str, Counter[str]] = defaultdict(Counter)
        theme_aliases: Dict[str, Counter[str]] = defaultdict(Counter)
        stock_buckets: Dict[str, Dict[str, Any]] = {}
        mention_dates: Dict[str, Counter[date]] = defaultdict(Counter)
        daily: Dict[date, Counter[str]] = defaultdict(Counter)
        evidence_count = verified_stock_row_count = falsification_count = 0

        def period_bucket(value: date) -> str:
            if granularity == "month":
                return value.strftime("%Y-%m")
            if granularity == "week":
                return (value - timedelta(days=value.weekday())).isoformat()
            return value.isoformat()

        # Old model outputs may contain a syntactically valid but factually wrong
        # code. Only expose a code when the local security master maps that exact
        # company name back to it; otherwise aggregate safely by company name.
        verified_name_to_code: Dict[str, str] = {}
        verified_code_to_name = get_stock_name_index_map()
        for code, indexed_name in verified_code_to_name.items():
            normalized_name = str(indexed_name or "").strip().casefold()
            normalized_code = str(code or "").strip().upper()
            if not normalized_name or not normalized_code:
                continue
            current = verified_name_to_code.get(normalized_name, "")
            if not current or ("." in normalized_code and "." not in current):
                verified_name_to_code[normalized_name] = normalized_code

        for row in rows:
            note = row.get("note") or {}
            source = str(note.get("group_name") or "未知来源").strip() or "未知来源"
            row_day = local_day(row)
            row_theme_pairs = canonicalize_topics(row.get("themes") or [])
            row_themes = list(dict.fromkeys(canonical for canonical, _raw in row_theme_pairs))
            for canonical, raw in row_theme_pairs:
                theme_aliases[canonical][raw] += 1
            row_mentions = []
            for mention in row.get("stock_mentions") or []:
                name = str(mention.get("name") or mention.get("ts_code") or "").strip()
                explicit_code = str(mention.get("ts_code") or "").strip().upper()
                verified_code = verified_name_to_code.get(name.casefold(), "")
                if explicit_code and str(verified_code_to_name.get(explicit_code) or "").strip().casefold() != name.casefold():
                    explicit_code = ""
                key = verified_code or explicit_code or name
                key = key.strip().upper()
                if not name or not key:
                    continue
                row_mentions.append((key, name, mention, verified_code or explicit_code))
            row_signals: List[tuple[str, str]] = []

            sources[source] += 1
            themes.update(row_themes)
            stocks.update(key for key, _, _, _ in row_mentions)
            signals.update(row_signals)
            for theme in row_themes:
                source_theme[(source, theme)] += 1
                if row_day:
                    theme_by_day[theme][period_bucket(row_day)] += 1
                for key, _, _, _ in row_mentions:
                    theme_stock[(theme, key)] += 1
            for key, name, mention, resolved_code in row_mentions:
                bucket = stock_buckets.setdefault(key, {
                    "ts_code": resolved_code, "name": name,
                    "current_count": 0, "previous_count": 0, "bullish": 0,
                    "bearish": 0, "neutral": 0, "watching": 0,
                    "importance_total": 0, "latest_at": None, "latest_thesis": None,
                })
                if resolved_code and not bucket["ts_code"]:
                    bucket["ts_code"] = resolved_code
                stance = str(mention.get("stance") or "neutral")
                if stance not in _STANCES:
                    stance = "neutral"
                bucket[stance] += 1
                bucket["importance_total"] += int(row.get("importance_score") or 0)
                created_at = note.get("created_at")
                if created_at and (not bucket["latest_at"] or created_at > bucket["latest_at"]):
                    bucket["latest_at"] = created_at
                    bucket["latest_thesis"] = row.get("summary")
                if row_day and row_day >= current_start:
                    bucket["current_count"] += 1
                elif row_day and previous_start <= row_day < current_start:
                    bucket["previous_count"] += 1
                if row_day and selected_start <= row_day <= selected_end:
                    mention_dates[key][row_day] += 1
                for kind, text in row_signals:
                    stock_signal[(key, kind, text)] += 1

            if row_day and row_day >= trend_start:
                sentiment = row.get("sentiment") if row.get("sentiment") in _SENTIMENTS else "neutral"
                daily[row_day]["total"] += 1
                daily[row_day][sentiment] += 1
            evidence_count += int(bool(row.get("evidence")))
            verified_stock_row_count += int(any(resolved_code for _, _, _, resolved_code in row_mentions))
            falsification_count += int(bool(row.get("falsification_conditions") or row.get("monitoring_points")))

        top_sources = {name for name, _ in sources.most_common(6)}
        top_themes = {name for name, _ in themes.most_common(7)}
        top_stocks = {name for name, _ in stocks.most_common(7)}
        market_impact = EssayMarketInsightService(self.repo.db).build(
            stock_buckets=stock_buckets,
            mention_dates=mention_dates,
            start_date=selected_start,
            end_date=selected_end,
            limit=8,
        )

        def nodes(counter: Counter, stage: str, limit: int) -> List[Dict[str, Any]]:
            result = []
            for raw, count in counter.most_common(limit):
                if isinstance(raw, tuple):
                    kind, label = raw
                    result.append({"stage": stage, "key": f"{kind}:{label}", "label": label, "kind": kind, "count": int(count)})
                else:
                    result.append({"stage": stage, "key": str(raw), "label": str(raw), "count": int(count)})
            return result

        stock_nodes = []
        for key, count in stocks.most_common(7):
            bucket = stock_buckets.get(key) or {}
            stock_nodes.append({
                "stage": "stocks", "key": key,
                "label": str(bucket.get("name") or key),
                "ts_code": str(bucket.get("ts_code") or ""),
                "count": int(count),
            })

        outcome_nodes = []
        for item in market_impact["items"]:
            if item["key"] not in top_stocks:
                continue
            metric = next((row for row in item["metrics"] if row["period"] == 5), None) or item["metrics"][0]
            average = metric.get("average_return")
            win_rate = metric.get("win_rate")
            if average is None:
                label = "5日样本未成熟"
                kind = "neutral"
            else:
                label = f"5日 {float(average):+.2f}% · 胜率 {float(win_rate or 0):.1f}%"
                kind = "positive" if float(average) > 0 else "negative" if float(average) < 0 else "neutral"
            outcome_nodes.append({
                "stage": "outcomes",
                "key": item["key"],
                "label": label,
                "stock_name": item["name"],
                "ts_code": item["ts_code"],
                "kind": kind,
                "count": int(metric.get("sample_count") or 0),
            })

        edges = [
            {"from_stage": "sources", "from": source, "to_stage": "themes", "to": theme, "count": int(count)}
            for (source, theme), count in source_theme.most_common()
            if source in top_sources and theme in top_themes
        ] + [
            {"from_stage": "themes", "from": theme, "to_stage": "stocks", "to": stock, "count": int(count)}
            for (theme, stock), count in theme_stock.most_common()
            if theme in top_themes and stock in top_stocks
        ] + [
            {"from_stage": "stocks", "from": item["key"], "to_stage": "outcomes", "to": item["key"], "count": int(item["covered_event_days"])}
            for item in market_impact["items"]
            if item["key"] in top_stocks
        ]

        if granularity == "day":
            dates = [(selected_start + timedelta(days=offset)).isoformat() for offset in range(safe_days)]
        elif granularity == "week":
            dates = []
            cursor = selected_start
            while cursor <= selected_end:
                bucket = period_bucket(cursor)
                if not dates or dates[-1] != bucket:
                    dates.append(bucket)
                cursor += timedelta(days=1)
        else:
            dates = []
            cursor = selected_start
            while cursor <= selected_end:
                bucket = period_bucket(cursor)
                if not dates or dates[-1] != bucket:
                    dates.append(bucket)
                cursor += timedelta(days=1)
        daily_theme_totals = {
            day: sum(theme_by_day[theme].get(day, 0) for theme in themes)
            for day in dates
        }
        daily_theme_peaks = {
            day: max((theme_by_day[theme].get(day, 0) for theme in themes), default=0)
            for day in dates
        }
        heatmap = [{
            "name": theme,
            "total": int(total),
            "aliases": [
                {"name": raw, "count": int(count)}
                for raw, count in theme_aliases[theme].most_common()
            ],
            "points": [{
                "date": day,
                "count": int(theme_by_day[theme].get(day, 0)),
                "daily_total": int(daily_theme_totals[day]),
                "share_percent": round(
                    theme_by_day[theme].get(day, 0) * 100 / daily_theme_totals[day], 2,
                ) if daily_theme_totals[day] else 0.0,
                "concentration_score": round(100 * (
                    (
                        theme_by_day[theme].get(day, 0) / daily_theme_totals[day]
                        * theme_by_day[theme].get(day, 0) / daily_theme_peaks[day]
                    ) ** 0.5
                ), 2) if daily_theme_totals[day] and daily_theme_peaks[day] else 0.0,
            } for day in dates],
        } for theme, total in themes.most_common(8)]

        momentum = []
        divergence = []
        for key, bucket in stock_buckets.items():
            total = bucket["bullish"] + bucket["bearish"] + bucket["neutral"] + bucket["watching"]
            current = bucket["current_count"]
            previous = bucket["previous_count"]
            change = current - previous
            row = {
                **{name: bucket[name] for name in ("ts_code", "name", "current_count", "previous_count", "bullish", "bearish", "neutral", "watching", "latest_at", "latest_thesis")},
                "change": change,
                "change_percent": round(change / previous * 100, 1) if previous else (100.0 if current else 0.0),
                "average_importance": round(bucket["importance_total"] / total, 1) if total else 0.0,
            }
            momentum.append(row)
            if bucket["bullish"] and bucket["bearish"]:
                divergence.append({
                    "key": key, "ts_code": bucket["ts_code"], "name": bucket["name"],
                    "bullish": bucket["bullish"], "bearish": bucket["bearish"],
                    "neutral": bucket["neutral"] + bucket["watching"], "total": total,
                    "divergence_score": round(200 * min(bucket["bullish"], bucket["bearish"]) / total, 1),
                })
        momentum.sort(key=lambda item: (item["current_count"], item["change"], item["average_importance"]), reverse=True)
        divergence.sort(key=lambda item: (item["divergence_score"], item["total"]), reverse=True)

        verification = [row for row in rows if (
            int(row.get("novelty_score") or 0) >= 70
            or float(row.get("confidence_score") or 0) < 0.55
            or bool(row.get("contradictions"))
        )]
        verification.sort(key=lambda row: (
            int(row.get("novelty_score") or 0),
            -float(row.get("confidence_score") or 0),
            int(row.get("importance_score") or 0),
        ), reverse=True)

        pulse: List[Dict[str, Any]] = []

        result = {
            "generated_at": utc_naive_now().isoformat() + "Z",
            "window_days": safe_days,
            "period": {
                "horizon": normalized_horizon,
                "start_date": selected_start.isoformat(),
                "end_date": selected_end.isoformat(),
                "granularity": granularity,
                "comparison_days": comparison_days,
            },
            "latest_data_at": max((str((row.get("note") or {}).get("created_at") or "") for row in rows), default=None),
            "summary": {
                "analyzed_count": len(rows), "source_count": len(sources),
                "theme_count": len(themes), "stock_count": len(stocks),
                "evidence_coverage_percent": round(evidence_count / len(rows) * 100, 1) if rows else 0.0,
                "high_novelty_count": sum(int(row.get("novelty_score") or 0) >= 70 for row in rows),
                "divergence_count": len(divergence),
            },
            "pulse": pulse,
            "layers": {
                "sources": nodes(sources, "sources", 6),
                "themes": nodes(themes, "themes", 7),
                "stocks": stock_nodes,
                "outcomes": outcome_nodes,
                "edges": edges,
            },
            "theme_heatmap": {
                "dates": dates,
                "items": heatmap,
                "taxonomy": {
                    "version": TOPIC_TAXONOMY_VERSION,
                    "raw_theme_count": len({raw for aliases in theme_aliases.values() for raw in aliases}),
                    "canonical_theme_count": len(themes),
                    "merged_theme_count": sum(max(0, len(aliases) - 1) for aliases in theme_aliases.values()),
                    "method": "强同义词与明确子主题归并；单篇同主题只计一次，原始标签完整保留",
                },
                "granularity": granularity,
            },
            "market_impact": market_impact,
            "stock_momentum": momentum[:12],
            "divergence": divergence[:10],
            "verification_queue": verification[:10],
            "evidence_funnel": [
                {"name": "已分析", "count": len(rows)},
                {"name": "含原文证据", "count": evidence_count},
                {"name": "匹配有效股票代码", "count": verified_stock_row_count},
                {"name": "含证伪或跟踪点", "count": falsification_count},
            ],
        }
        with _deep_insights_cache_lock:
            stored_at = time.monotonic()
            _deep_insights_cache[cache_key] = (stored_at, result)
            _prune_ttl_cache(
                _deep_insights_cache,
                now=stored_at,
                ttl=_DEEP_INSIGHTS_CACHE_TTL_SECONDS,
                max_entries=_DEEP_INSIGHTS_CACHE_MAX_ENTRIES,
            )
        return result

    def interpret_market_impact(
        self,
        *,
        ts_code: str,
        horizon: str,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Explain an already-computed market study; never ask the model to invent data."""
        atlas = self.deep_insights(
            horizon=horizon,
            start_date=start_date,
            end_date=end_date,
        )
        normalized = str(ts_code or "").strip().upper()
        item = next((
            row for row in atlas["market_impact"]["items"]
            if str(row.get("ts_code") or "").strip().upper() == normalized
        ), None)
        if item is None:
            raise KeyError("当前研究窗口没有该股票的可用行情关联样本")
        analyzer = DeepSeekEssayAnalyzer(timeout_seconds=90)
        if not analyzer.configured:
            raise EssayAnalysisError("DeepSeek 尚未配置，无法生成行情关联解读")
        context = {
            "period": atlas["period"],
            "stock": {
                key: item[key]
                for key in (
                    "ts_code", "name", "mention_count", "event_day_count",
                    "covered_event_days", "metrics", "lead_lag", "latest_price_date",
                    "latest_close", "data_source",
                )
            },
            "method": {
                key: atlas["market_impact"][key]
                for key in ("benchmark", "entry_rule", "exit_rule", "price_basis", "dedupe_rule", "causality_note")
            },
        }
        payload = {
            "model": analyzer.model,
            "messages": [
                {
                    "role": "system",
                    "content": (
                        "你是严谨的事件研究解释员。只能解释输入JSON中的统计结果，不得补造新闻、公司事实、价格或因果关系。"
                        "必须区分相关性与因果；样本少于10时明确提示小样本。输出JSON对象："
                        '{"conclusion":"不超过100字","evidence":["最多3条"],"limitations":["最多3条"],'
                        '"next_checks":["最多3条"]}'
                    ),
                },
                {"role": "user", "content": json.dumps(context, ensure_ascii=False)},
            ],
            "response_format": {"type": "json_object"},
            "thinking": {"type": "disabled"},
            "temperature": 0.1,
            "max_tokens": 1200,
            "stream": False,
        }
        response = analyzer._post_with_retry(payload)
        parsed = analyzer._parse_json(analyzer._extract_content(response))
        return {
            "generated_at": utc_naive_now().isoformat() + "Z",
            "model": analyzer.model,
            "ts_code": item["ts_code"],
            "period": atlas["period"],
            "interpretation": {
                "conclusion": str(parsed.get("conclusion") or "")[:500],
                "evidence": [str(value)[:500] for value in (parsed.get("evidence") or [])[:3]],
                "limitations": [str(value)[:500] for value in (parsed.get("limitations") or [])[:3]],
                "next_checks": [str(value)[:500] for value in (parsed.get("next_checks") or [])[:3]],
            },
        }

    @staticmethod
    def _configured_watchlist() -> List[tuple[str, str]]:
        return EssayDailyReportService._daily_watchlist()

    @staticmethod
    def _compare_reports(rows: Sequence[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
        """Compare structured theme/stock directions across completed model reports."""
        if len(rows) < 2:
            return {"consensus": [], "divergences": []}
        direction_labels = {
            "bullish": "看多", "bearish": "看空", "mixed": "分歧", "neutral": "中性",
            "watching": "观察",
        }
        themes: Dict[str, Dict[str, Any]] = {}
        stocks: Dict[str, Dict[str, Any]] = {}
        exact_consensus, exact_divergences = Counter(), Counter()
        for row in rows:
            model = str(row.get("model") or "unknown")
            report = row.get("report") or {}
            exact_consensus.update(str(value).strip() for value in report.get("consensus") or [] if str(value).strip())
            exact_divergences.update(str(value).strip() for value in report.get("divergences") or [] if str(value).strip())
            for item in report.get("key_themes") or []:
                if not isinstance(item, dict):
                    continue
                name = str(item.get("name") or "").strip()
                if not name:
                    continue
                bucket = themes.setdefault(name.lower(), {"name": name, "models": set(), "directions": Counter()})
                bucket["models"].add(model)
                bucket["directions"][str(item.get("direction") or "neutral").lower()] += 1
            for item in report.get("stock_focus") or []:
                if not isinstance(item, dict):
                    continue
                name = str(item.get("name") or item.get("ts_code") or "").strip()
                key = str(item.get("ts_code") or name).strip().upper()
                if not key:
                    continue
                bucket = stocks.setdefault(key, {"name": name, "models": set(), "directions": Counter()})
                bucket["models"].add(model)
                bucket["directions"][str(item.get("stance") or "neutral").lower()] += 1

        consensus: List[Dict[str, Any]] = []
        divergences: List[Dict[str, Any]] = []
        for kind, buckets in (("主题", themes), ("股票", stocks)):
            for bucket in buckets.values():
                model_count = len(bucket["models"])
                if model_count < 2:
                    continue
                directions = bucket["directions"]
                meaningful = {key for key, count in directions.items() if count and key not in {"neutral", "watching"}}
                direction_text = "、".join(
                    f"{direction_labels.get(key, key)} {count}个模型" for key, count in directions.most_common()
                )
                entry = {"text": f"{kind}「{bucket['name']}」：{direction_text}", "model_count": model_count}
                if len(meaningful) > 1 or "mixed" in meaningful:
                    divergences.append(entry)
                else:
                    consensus.append(entry)
        consensus.extend(
            {"text": text, "model_count": count}
            for text, count in exact_consensus.most_common() if count >= 2
        )
        divergences.extend(
            {"text": text, "model_count": count}
            for text, count in exact_divergences.most_common() if count >= 2
        )

        def dedupe(items: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
            result, seen = [], set()
            for item in sorted(items, key=lambda value: (-int(value["model_count"]), value["text"])):
                if item["text"] in seen:
                    continue
                seen.add(item["text"])
                result.append(item)
            return result[:12]
        return {"consensus": dedupe(consensus), "divergences": dedupe(divergences)}

    @staticmethod
    def _watchlist_summary(symbol: str, name: str, rows: Sequence[Dict[str, Any]], local_day, today: date, trend_days: int) -> Dict[str, Any]:
        sorted_rows = sorted(rows, key=lambda row: str((row.get("note") or {}).get("created_at") or ""), reverse=True)
        stances, timeline = Counter(), defaultdict(Counter)
        for row in rows:
            day = local_day(row)
            if day:
                timeline[day]["total"] += 1
            for mention in row.get("stock_mentions") or []:
                if symbol.upper() == str(mention.get("ts_code") or "").upper() or name.lower() in str(mention.get("name") or "").lower():
                    stance = mention.get("stance") if mention.get("stance") in _STANCES else "neutral"
                    stances[stance] += 1
                    if day:
                        timeline[day][stance] += 1
        def period_count(days: int) -> int:
            start = today - timedelta(days=days - 1)
            return sum(1 for row in rows if (local_day(row) or date.min) >= start)
        latest = sorted_rows[0] if sorted_rows else None
        trend_start = today - timedelta(days=trend_days - 1)
        return {
            "symbol": symbol, "name": name, "mention_count": len(rows),
            "day_mentions": period_count(1), "week_mentions": period_count(7), "month_mentions": period_count(30),
            "stances": dict(stances),
            "average_importance": round(sum(int(row.get("importance_score") or 0) for row in rows) / len(rows), 1) if rows else 0,
            "average_confidence": round(sum(float(row.get("confidence_score") or 0) for row in rows) / len(rows), 3) if rows else 0,
            "latest_at": (latest.get("note") or {}).get("created_at") if latest else None,
            "latest_thesis": latest.get("summary") if latest else None,
            "catalysts": list(dict.fromkeys(value for row in sorted_rows[:10] for value in (row.get("catalysts") or [])))[:6],
            "risks": list(dict.fromkeys(value for row in sorted_rows[:10] for value in (row.get("risks") or [])))[:6],
            "trend": [{"date": (trend_start + timedelta(days=offset)).isoformat(), **dict(timeline[trend_start + timedelta(days=offset)])} for offset in range(trend_days)],
            "latest_items": sorted_rows[:8],
        }

    def word_cloud(
        self,
        *,
        period: str = "day",
        anchor_date: Optional[str] = None,
        kind: str = "stocks",
        stock: Optional[str] = None,
        top_n: int = 80,
    ) -> Dict[str, Any]:
        if period not in {"day", "week", "month"}:
            raise ValueError("period must be day, week or month")
        if kind not in {"stocks", "tags", "themes"}:
            raise ValueError("kind must be stocks, tags or themes")
        anchor = date.fromisoformat(anchor_date) if anchor_date else datetime.now(timezone(timedelta(hours=8))).date()
        if period == "day":
            local_start, local_end = anchor, anchor + timedelta(days=1)
        elif period == "week":
            local_start = anchor - timedelta(days=anchor.weekday())
            local_end = local_start + timedelta(days=7)
        else:
            local_start = anchor.replace(day=1)
            local_end = (local_start.replace(day=28) + timedelta(days=4)).replace(day=1)
        start, _ = EssayDailyReportService._utc_bounds(local_start)
        actual_end, _ = EssayDailyReportService._utc_bounds(local_end)
        window = actual_end - start
        rows = self.repo.completed_between(start=start, end=actual_end)
        previous = self.repo.completed_between(start=start - window, end=start)
        current_counts, metadata = self._cloud_counts(rows, kind=kind, stock=stock)
        previous_counts, _ = self._cloud_counts(previous, kind=kind, stock=stock)
        items = []
        for name, count in current_counts.most_common(max(10, min(int(top_n), 200))):
            prev = int(previous_counts.get(name, 0))
            item = {"name": name, "count": int(count), "previous_count": prev, "change": int(count) - prev}
            item.update(metadata.get(name, {}))
            items.append(item)
        return {
            "period": period,
            "kind": kind,
            "stock": stock,
            "start_date": local_start.isoformat(),
            "end_date": (local_end - timedelta(days=1)).isoformat(),
            "source_count": len(rows),
            "items": items,
        }

    @staticmethod
    def _cloud_counts(rows: Sequence[Dict[str, Any]], *, kind: str, stock: Optional[str]) -> tuple[Counter, Dict[str, Dict[str, Any]]]:
        counts: Counter = Counter()
        metadata: Dict[str, Dict[str, Any]] = {}
        stock_query = str(stock or "").strip().lower()
        for row in rows:
            mentions = row.get("stock_mentions") or []
            if stock_query and not any(stock_query in str(item.get("ts_code") or "").lower() or stock_query in str(item.get("name") or "").lower() for item in mentions):
                continue
            if kind == "stocks":
                for item in mentions:
                    name = str(item.get("name") or item.get("ts_code") or "").strip()
                    if not name or name in {"未知", "不详", "公司", "标的"}:
                        continue
                    counts[name] += 1
                    meta = metadata.setdefault(name, {"ts_code": item.get("ts_code") or "", "bullish": 0, "bearish": 0, "neutral": 0, "watching": 0})
                    stance = item.get("stance") if item.get("stance") in _STANCES else "neutral"
                    meta[stance] += 1
            else:
                counts.update(row.get(kind) or [])
        return counts, metadata

    @staticmethod
    def _counter_rows(counter: Counter, limit: Optional[int] = None) -> List[Dict[str, Any]]:
        rows = counter.most_common(limit)
        return [{"name": str(name), "count": int(count)} for name, count in rows]
