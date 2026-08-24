# -*- coding: utf-8 -*-
"""Local, human-filterable broker research-report library."""

from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, timedelta
from hashlib import sha256
from io import BytesIO
import json
import logging
import threading
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import and_, asc, case, desc, func, or_, select

from src.services.financial_data_service import FinancialDataService
from src.storage import (
    DatabaseManager,
    ResearchReportRecord,
    ResearchReportSyncState,
    utc_naive_now,
)

logger = logging.getLogger(__name__)
_SOURCE_KEY = "tushare.research_report"
_FIELDS = [
    "trade_date", "abstr", "title", "report_type", "author", "name",
    "ts_code", "inst_csname", "ind_name", "url",
]
_DEEP_TERMS = ("深度", "专题", "全景", "产业链", "白皮书", "年度策略", "投资策略")
_QUICK_TERMS = ("日报", "周报", "早报", "晨报", "快评", "点评")


class ResearchReportLibraryError(RuntimeError):
    """Safe error exposed by the report-library API."""


class ResearchReportLibraryService:
    """Backfill Tushare report metadata once, then search SQLite only."""

    _instance: Optional["ResearchReportLibraryService"] = None
    _instance_lock = threading.Lock()

    def __init__(
        self,
        *,
        db_manager: Optional[DatabaseManager] = None,
        financial: Optional[FinancialDataService] = None,
    ) -> None:
        self.db = db_manager or DatabaseManager.get_instance()
        self.financial = financial or FinancialDataService()
        self._executor = ThreadPoolExecutor(max_workers=1, thread_name_prefix="report-library")
        self._run_lock = threading.RLock()
        self._future = None
        self._repair_interrupted_state()

    @classmethod
    def get_instance(cls) -> "ResearchReportLibraryService":
        with cls._instance_lock:
            if cls._instance is None:
                cls._instance = cls()
            return cls._instance

    @classmethod
    def reset_instance(cls) -> None:
        with cls._instance_lock:
            if cls._instance is not None:
                cls._instance._executor.shutdown(wait=False, cancel_futures=True)
            cls._instance = None

    def ensure_background_sync(self) -> Dict[str, Any]:
        """Start the two-year backfill or a short incremental refresh when due."""
        status = self.status()
        if status["status"] in {"queued", "running"}:
            return status
        today = date.today()
        earliest = self._parse_date(status.get("earliest_trade_date"))
        latest = self._parse_date(status.get("latest_trade_date"))
        completed = self._parse_datetime(status.get("completed_at"))
        due = completed is None or utc_naive_now() - completed >= timedelta(hours=6)
        target_start = today - timedelta(days=730)
        # The exact boundary can be a weekend/holiday with no published report.
        needs_two_years = not status["total"] or earliest is None or earliest > target_start + timedelta(days=14)
        needs_increment = latest is None or latest < today - timedelta(days=2)
        if needs_two_years:
            return self.start_sync(start_date=today - timedelta(days=730), end_date=today)
        if due and needs_increment:
            return self.start_sync(start_date=today - timedelta(days=14), end_date=today)
        return status

    def start_sync(
        self,
        *,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        force: bool = False,
    ) -> Dict[str, Any]:
        with self._run_lock:
            if self._future is not None and not self._future.done():
                return self.status()
            current = self.status()
            if current["status"] in {"queued", "running"} and not force:
                return current
            end = end_date or date.today()
            start = start_date or (end - timedelta(days=730))
            if start > end:
                raise ResearchReportLibraryError("研报同步开始日期不能晚于结束日期")
            windows = self._date_windows(start, end)
            self._write_state(
                status="queued",
                progress=0,
                message="过去两年研报链接已进入后台同步队列",
                start_date=start,
                end_date=end,
                total_windows=len(windows),
                completed_windows=0,
                scanned_rows=0,
                saved_rows=0,
                last_error=None,
                started_at=utc_naive_now(),
                completed_at=None,
            )
            self._future = self._executor.submit(self._run_sync, start, end, windows)
            return self.status()

    def status(self) -> Dict[str, Any]:
        with self.db.get_session() as session:
            state = session.get(ResearchReportSyncState, _SOURCE_KEY)
            total, earliest, latest, pdf_count = session.execute(select(
                func.count(ResearchReportRecord.id),
                func.min(ResearchReportRecord.trade_date),
                func.max(ResearchReportRecord.trade_date),
                func.sum(case((ResearchReportRecord.pdf_url.is_not(None), 1), else_=0)),
            )).one()
        payload = {
            "status": state.status if state else "idle",
            "progress": int(state.progress or 0) if state else 0,
            "message": state.message if state else "等待首次同步",
            "start_date": self._date_text(state.start_date) if state else None,
            "end_date": self._date_text(state.end_date) if state else None,
            "total_windows": int(state.total_windows or 0) if state else 0,
            "completed_windows": int(state.completed_windows or 0) if state else 0,
            "scanned_rows": int(state.scanned_rows or 0) if state else 0,
            "saved_rows": int(state.saved_rows or 0) if state else 0,
            "last_error": state.last_error if state else None,
            "started_at": self._datetime_text(state.started_at) if state else None,
            "completed_at": self._datetime_text(state.completed_at) if state else None,
            "total": int(total or 0),
            "pdf_count": int(pdf_count or 0),
            "earliest_trade_date": self._date_text(earliest),
            "latest_trade_date": self._date_text(latest),
            "source": _SOURCE_KEY,
            "search_mode": "local_sqlite_only",
        }
        return payload

    def facets(self) -> Dict[str, Any]:
        return {
            "brokers": self._facet(ResearchReportRecord.broker, 200),
            "report_types": self._facet(ResearchReportRecord.report_type, 50),
            "industries": self._facet(ResearchReportRecord.industry, 300),
            "companies": self._facet(ResearchReportRecord.company_name, 500),
            "tags": self._tag_facets(),
        }

    def search(self, **filters: Any) -> Dict[str, Any]:
        page = max(1, int(filters.get("page") or 1))
        page_size = max(1, min(int(filters.get("page_size") or 30), 100))
        conditions = []
        self._append_text_conditions(conditions, ResearchReportRecord.title, filters.get("title_query"))
        self._append_text_conditions(conditions, ResearchReportRecord.abstract, filters.get("content_query"))
        for key, column in (
            ("broker", ResearchReportRecord.broker),
            ("report_type", ResearchReportRecord.report_type),
            ("industry", ResearchReportRecord.industry),
            ("ts_code", ResearchReportRecord.ts_code),
        ):
            value = str(filters.get(key) or "").strip()
            if value:
                conditions.append(column == value)
        for key, columns in (
            ("company", (ResearchReportRecord.company_name, ResearchReportRecord.ts_code)),
            ("author", (ResearchReportRecord.author,)),
        ):
            value = str(filters.get(key) or "").strip()
            if value:
                pattern = self._like(value)
                conditions.append(or_(*(column.like(pattern, escape="\\") for column in columns)))
        tag = str(filters.get("tag") or "").strip()
        if tag:
            conditions.append(ResearchReportRecord.tags_json.like(self._like(f'"{tag}"'), escape="\\"))
        start = self._parse_date(filters.get("start_date"))
        end = self._parse_date(filters.get("end_date"))
        if start:
            conditions.append(ResearchReportRecord.trade_date >= start)
        if end:
            conditions.append(ResearchReportRecord.trade_date <= end)
        if self._as_bool(filters.get("has_pdf"), default=True):
            conditions.extend((ResearchReportRecord.pdf_url.is_not(None), ResearchReportRecord.pdf_url != ""))
        where = and_(*conditions) if conditions else True
        order = asc(ResearchReportRecord.trade_date) if filters.get("sort") == "oldest" else desc(ResearchReportRecord.trade_date)
        with self.db.get_session() as session:
            total = session.execute(select(func.count(ResearchReportRecord.id)).where(where)).scalar() or 0
            rows = session.execute(
                select(ResearchReportRecord).where(where).order_by(order, desc(ResearchReportRecord.id))
                .offset((page - 1) * page_size).limit(page_size)
            ).scalars().all()
        return {
            "items": [self._serialize(row) for row in rows],
            "total": int(total),
            "page": page,
            "page_size": page_size,
            "source": "local_sqlite",
        }

    def export_selected(self, ids: Iterable[int]) -> bytes:
        selected_ids = list(dict.fromkeys(int(value) for value in ids))[:1000]
        if not selected_ids:
            raise ResearchReportLibraryError("请至少选择一篇研报")
        with self.db.get_session() as session:
            rows = session.execute(
                select(ResearchReportRecord).where(ResearchReportRecord.id.in_(selected_ids))
                .order_by(desc(ResearchReportRecord.trade_date), desc(ResearchReportRecord.id))
            ).scalars().all()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "已选研报"
        headers = ["日期", "标题", "摘要", "研报类型", "券商", "作者", "公司", "股票代码", "行业", "人工筛选标签", "PDF链接"]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for row in rows:
            values = self._serialize(row)
            sheet.append([
                values["trade_date"], values["title"], values["abstract"], values["report_type"],
                values["broker"], values["author"], values["company_name"], values["ts_code"],
                values["industry"], " / ".join(values["tags"]), values["pdf_url"],
            ])
            if values["pdf_url"]:
                cell = sheet.cell(row=sheet.max_row, column=11)
                cell.hyperlink = values["pdf_url"]
                cell.style = "Hyperlink"
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column, width in {"A": 13, "B": 55, "C": 70, "D": 14, "E": 16, "F": 20, "G": 18, "H": 14, "I": 16, "J": 30, "K": 50}.items():
            sheet.column_dimensions[column].width = width
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _run_sync(self, start: date, end: date, windows: List[tuple[date, date]]) -> None:
        scanned = 0
        saved = 0
        try:
            self._write_state(status="running", progress=1, message="正在从 Tushare 分页同步研报元数据和 PDF 链接")
            for index, (window_start, window_end) in enumerate(windows, start=1):
                offset = 0
                while True:
                    result = self.financial.query(
                        source="tushare",
                        resource="research_report",
                        params={
                            "start_date": window_start.strftime("%Y%m%d"),
                            "end_date": window_end.strftime("%Y%m%d"),
                            "limit": 1000,
                            "offset": offset,
                        },
                        fields=_FIELDS,
                    )
                    rows = list(result.get("rows") or [])
                    scanned += len(rows)
                    saved += self._upsert(rows)
                    if len(rows) < 1000:
                        break
                    offset += 1000
                    if offset >= 50_000:
                        raise ResearchReportLibraryError("单个日期窗口超过安全分页上限，请缩短同步窗口")
                progress = min(99, max(1, round(index / max(1, len(windows)) * 99)))
                self._write_state(
                    status="running",
                    progress=progress,
                    completed_windows=index,
                    scanned_rows=scanned,
                    saved_rows=saved,
                    message=f"研报链接同步 {index}/{len(windows)} · 已扫描 {scanned:,} 条",
                )
            self._write_state(
                status="completed",
                progress=100,
                completed_windows=len(windows),
                scanned_rows=scanned,
                saved_rows=saved,
                message=f"研报库同步完成：扫描 {scanned:,} 条，新增或更新 {saved:,} 条",
                completed_at=utc_naive_now(),
                last_error=None,
            )
        except Exception as exc:  # noqa: BLE001 - failure must become durable status.
            logger.exception("Research report library sync failed")
            self._write_state(
                status="failed",
                progress=min(99, int(self.status().get("progress") or 0)),
                scanned_rows=scanned,
                saved_rows=saved,
                message=f"研报链接同步失败：{type(exc).__name__}",
                last_error=str(exc)[:500],
                completed_at=utc_naive_now(),
            )

    def _upsert(self, raw_rows: List[Dict[str, Any]]) -> int:
        normalized = [self._normalize(row) for row in raw_rows]
        normalized = [row for row in normalized if row is not None]
        if not normalized:
            return 0
        keys = [row["report_key"] for row in normalized]
        changed = 0
        with self.db.get_session() as session:
            existing = session.execute(
                select(ResearchReportRecord).where(ResearchReportRecord.report_key.in_(keys))
            ).scalars().all()
            by_key = {row.report_key: row for row in existing}
            for fields in normalized:
                record = by_key.get(fields["report_key"])
                if record is None:
                    session.add(ResearchReportRecord(**fields))
                    changed += 1
                    continue
                for key, value in fields.items():
                    if key != "report_key":
                        setattr(record, key, value)
                changed += 1
            session.commit()
        return changed

    @staticmethod
    def _normalize(row: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        title = str(row.get("title") or "").strip()
        trade_date = ResearchReportLibraryService._parse_date(row.get("trade_date"))
        if not title or trade_date is None:
            return None
        pdf_url = str(row.get("url") or "").strip()
        report_key = sha256(
            json.dumps([trade_date.isoformat(), title, pdf_url], ensure_ascii=False).encode("utf-8")
        ).hexdigest()
        report_type = str(row.get("report_type") or "").strip()
        broker = str(row.get("inst_csname") or "").strip()
        industry = str(row.get("ind_name") or "").strip()
        company = str(row.get("name") or "").strip()
        tags = [value for value in (report_type, industry, broker, company) if value]
        if any(term in title for term in _DEEP_TERMS):
            tags.append("深度研究")
        if any(term in title for term in _QUICK_TERMS):
            tags.append("快评跟踪")
        if pdf_url:
            tags.append("有PDF")
        return {
            "report_key": report_key,
            "trade_date": trade_date,
            "title": title[:800],
            "abstract": str(row.get("abstr") or "").strip(),
            "report_type": report_type[:80] or None,
            "author": str(row.get("author") or "").strip()[:300] or None,
            "company_name": company[:160] or None,
            "ts_code": str(row.get("ts_code") or "").strip()[:20] or None,
            "broker": broker[:160] or None,
            "industry": industry[:160] or None,
            "pdf_url": pdf_url or None,
            "tags_json": json.dumps(list(dict.fromkeys(tags)), ensure_ascii=False),
            "source": _SOURCE_KEY,
            "synced_at": utc_naive_now(),
        }

    def _write_state(self, **changes: Any) -> None:
        with self.db.get_session() as session:
            state = session.get(ResearchReportSyncState, _SOURCE_KEY)
            if state is None:
                state = ResearchReportSyncState(source_key=_SOURCE_KEY)
                session.add(state)
            for key, value in changes.items():
                if hasattr(state, key):
                    setattr(state, key, value)
            state.updated_at = utc_naive_now()
            session.commit()

    def _repair_interrupted_state(self) -> None:
        with self.db.get_session() as session:
            state = session.get(ResearchReportSyncState, _SOURCE_KEY)
            if state is None or state.status not in {"queued", "running"}:
                return
            state.status = "interrupted"
            state.message = "服务重启中断了研报同步，系统将自动续跑"
            state.last_error = "service_restarted"
            state.updated_at = utc_naive_now()
            session.commit()

    def _facet(self, column, limit: int) -> List[Dict[str, Any]]:
        with self.db.get_session() as session:
            rows = session.execute(
                select(column, func.count(ResearchReportRecord.id))
                .where(column.is_not(None), column != "")
                .group_by(column).order_by(desc(func.count(ResearchReportRecord.id)), asc(column)).limit(limit)
            ).all()
        return [{"value": str(value), "count": int(count)} for value, count in rows]

    def _tag_facets(self) -> List[Dict[str, Any]]:
        counts: Dict[str, int] = {}
        with self.db.get_session() as session:
            rows = session.execute(select(ResearchReportRecord.tags_json)).scalars().all()
        for raw in rows:
            try:
                values = json.loads(raw or "[]")
            except (TypeError, ValueError):
                values = []
            for value in values:
                counts[str(value)] = counts.get(str(value), 0) + 1
        preferred = {"深度研究", "快评跟踪", "有PDF", "个股研报", "行业研报", "策略研报"}
        items = [(key, value) for key, value in counts.items() if key in preferred]
        return [{"value": key, "count": value} for key, value in sorted(items, key=lambda item: (-item[1], item[0]))]

    @staticmethod
    def _append_text_conditions(conditions: List[Any], column, value: Any) -> None:
        for term in str(value or "").split():
            conditions.append(column.like(ResearchReportLibraryService._like(term), escape="\\"))

    @staticmethod
    def _like(value: str) -> str:
        escaped = str(value).replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
        return f"%{escaped}%"

    @staticmethod
    def _date_windows(start: date, end: date) -> List[tuple[date, date]]:
        windows: List[tuple[date, date]] = []
        cursor = start
        while cursor <= end:
            window_end = min(end, cursor + timedelta(days=13))
            windows.append((cursor, window_end))
            cursor = window_end + timedelta(days=1)
        return windows

    @staticmethod
    def _parse_date(value: Any) -> Optional[date]:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        raw = str(value or "").strip()
        if not raw:
            return None
        for fmt in ("%Y-%m-%d", "%Y%m%d"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def _parse_datetime(value: Any) -> Optional[datetime]:
        if isinstance(value, datetime):
            return value
        raw = str(value or "").strip()
        if not raw:
            return None
        try:
            return datetime.fromisoformat(raw.replace("Z", "+00:00")).replace(tzinfo=None)
        except ValueError:
            return None

    @staticmethod
    def _as_bool(value: Any, *, default: bool) -> bool:
        if value in (None, ""):
            return default
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() not in {"0", "false", "no", "off"}

    @staticmethod
    def _serialize(row: ResearchReportRecord) -> Dict[str, Any]:
        try:
            tags = json.loads(row.tags_json or "[]")
        except (TypeError, ValueError):
            tags = []
        return {
            "id": row.id,
            "trade_date": row.trade_date.isoformat() if row.trade_date else None,
            "title": row.title,
            "abstract": row.abstract or "",
            "report_type": row.report_type,
            "author": row.author,
            "company_name": row.company_name,
            "ts_code": row.ts_code,
            "broker": row.broker,
            "industry": row.industry,
            "pdf_url": row.pdf_url,
            "tags": tags if isinstance(tags, list) else [],
            "synced_at": ResearchReportLibraryService._datetime_text(row.synced_at),
        }

    @staticmethod
    def _date_text(value: Any) -> Optional[str]:
        return value.isoformat() if isinstance(value, date) else None

    @staticmethod
    def _datetime_text(value: Any) -> Optional[str]:
        return f"{value.isoformat()}Z" if isinstance(value, datetime) else None
