# -*- coding: utf-8 -*-
"""Durable queue, DeepSeek normalization and essay-radar API tests."""

from __future__ import annotations

from io import BytesIO
import os
from datetime import date
from unittest.mock import MagicMock

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from api.v1.endpoints import essay_radar
from src.config import Config
from src.repositories.essay_analysis_repo import EssayAnalysisRepository, _select_in_batches
from src.services.essay_analysis_service import (
    DeepSeekEssayAnalyzer,
    EssayAnalysisService,
    EssayDailyReportService,
    _dashboard_rows_cache,
    _deep_insights_cache,
)
from src.services.essay_analysis_worker import EssayAnalysisWorker
import src.services.essay_analysis_worker as essay_worker_module
from src.services.financial_data_service import ResearchNoteService
from src.storage import DatabaseManager, EssayAnalysisRecord, ResearchNote, StockDaily
from src.utils.essay_topic_taxonomy import canonicalize_topic, canonicalize_topics, topic_search_terms


@pytest.fixture()
def essay_service(tmp_path):
    old_database_path = os.environ.get("DATABASE_PATH")
    os.environ["DATABASE_PATH"] = str(tmp_path / "essay-radar.db")
    Config.reset_instance()
    DatabaseManager.reset_instance()
    _dashboard_rows_cache.clear()
    _deep_insights_cache.clear()
    notes = ResearchNoteService()
    notes.import_topics([{
        "topic_id": "topic-1",
        "title": "机器人订单增长",
        "content": "公司称机器人订单增长，量产进度仍存在风险。",
        "create_time": "2026-08-19T14:26:11+0800",
        "group": {"group_id": "g1", "name": "调研纪要"},
    }])
    service = EssayAnalysisService()
    try:
        yield service
    finally:
        _dashboard_rows_cache.clear()
        _deep_insights_cache.clear()
        DatabaseManager.reset_instance()
        Config.reset_instance()
        if old_database_path is None:
            os.environ.pop("DATABASE_PATH", None)
        else:
            os.environ["DATABASE_PATH"] = old_database_path


def test_analyzer_normalizes_json_output_and_stock_codes(monkeypatch) -> None:
    response = MagicMock(status_code=200)
    response.json.return_value = {
        "choices": [{"message": {"content": """{
          "items": [{
            "topic_id": "topic-1", "summary": "订单增长但量产有风险",
            "primary_category": "company_research", "sentiment": "mixed",
            "time_horizon": "medium", "importance_score": 83,
            "confidence_score": 0.88, "tags": ["订单增长", "风险提示"],
            "industries": ["机械设备"], "themes": ["机器人"],
            "stock_mentions": [{"ts_code": "688836.SS", "name": "宇树科技",
              "stance": "watching", "confidence": 0.8, "rationale": "等待量产"}],
            "key_points": ["订单增长"], "catalysts": ["量产"], "risks": ["交付延期"]
          }]
        }"""}}],
        "usage": {"prompt_tokens": 100, "completion_tokens": 50, "total_tokens": 150},
    }
    session = MagicMock()
    session.post.return_value = response
    analyzer = DeepSeekEssayAnalyzer(api_key="test-key", session=session)

    result = analyzer.analyze_batch([{
        "topic_id": "topic-1",
        "title": "机器人订单增长",
        "content": "正文",
        "existing_symbols": [],
    }])

    assert result["items"][0]["stock_mentions"][0]["ts_code"] == "688836.SH"
    assert result["items"][0]["importance_score"] == 83
    assert result["usage"]["total_tokens"] == 150
    payload = session.post.call_args.kwargs["json"]
    assert payload["response_format"] == {"type": "json_object"}
    assert payload["thinking"] == {"type": "disabled"}


def test_analyzer_rejects_placeholder_summary_for_retry() -> None:
    response = MagicMock(status_code=200)
    response.json.return_value = {
        "choices": [{"message": {"content": """{
          "items": [{
            "topic_id": "topic-1",
            "summary": "信息不足，未形成有效摘要",
            "primary_category": "other",
            "sentiment": "neutral",
            "time_horizon": "unclear"
          }]
        }"""}}],
        "usage": {},
    }
    session = MagicMock()
    session.post.return_value = response
    analyzer = DeepSeekEssayAnalyzer(api_key="test-key", session=session)

    result = analyzer.analyze_batch([{
        "topic_id": "topic-1",
        "title": "只有标题也应形成摘要",
        "content": "",
        "existing_symbols": [],
    }])

    assert result["items"] == []


def test_worker_retries_only_missing_batch_item_as_single(monkeypatch) -> None:
    repository = MagicMock()
    repository.save_successes.side_effect = [1, 1]
    analyzer = MagicMock()
    analyzer.analyze_batch.side_effect = [
        {
            "items": [{"topic_id": "topic-good"}],
            "raw_response": "batch",
            "usage": {},
        },
        {
            "items": [{"topic_id": "topic-retry"}],
            "raw_response": "single",
            "usage": {},
        },
    ]
    monkeypatch.setattr(essay_worker_module, "EssayAnalysisRepository", lambda: repository)
    monkeypatch.setattr(essay_worker_module, "DeepSeekEssayAnalyzer", lambda: analyzer)

    result = EssayAnalysisWorker()._process_batch([
        {"topic_id": "topic-good"},
        {"topic_id": "topic-retry"},
    ])

    assert [len(call.args[0]) for call in analyzer.analyze_batch.call_args_list] == [2, 1]
    assert result == {"saved": 2, "failed": 0, "error": None}
    repository.save_failures.assert_not_called()


def test_topic_taxonomy_merges_optical_communication_synonyms_without_duplicate_record_hits() -> None:
    assert canonicalize_topic("CPO") == "光通信"
    assert canonicalize_topic("npo 光模块") == "光通信"
    assert canonicalize_topic("硅光子") == "光通信"
    assert canonicalize_topic("机器人") == "机器人"
    assert {canonical for canonical, _raw in canonicalize_topics(["CPO", "NPO", "光模块"])} == {"光通信"}
    assert {"光通信", "CPO", "NPO", "光模块"}.issubset(set(topic_search_terms("光通信")))


def test_deep_insights_merges_raw_topics_and_expands_feed_search(essay_service) -> None:
    ResearchNoteService().import_topics([{
        "topic_id": "topic-cpo",
        "title": "CPO 与光模块更新",
        "content": "CPO 光模块需求更新",
        "create_time": "2026-08-20T10:00:00+0800",
        "group": {"group_id": "g1", "name": "调研纪要"},
    }, {
        "topic_id": "topic-npo",
        "title": "NPO 路线更新",
        "content": "NPO 进入验证阶段",
        "create_time": "2026-08-20T11:00:00+0800",
        "group": {"group_id": "g1", "name": "调研纪要"},
    }])
    claimed = essay_service.repo.claim_batch(limit=10, max_attempts=4)
    themes = {
        "topic-1": ["CPO", "光模块"],
        "topic-cpo": ["光模块"],
        "topic-npo": ["NPO"],
    }
    essay_service.repo.save_successes([{
        "topic_id": item["topic_id"],
        "summary": "光通信产业链更新",
        "primary_category": "industry_chain",
        "sentiment": "neutral",
        "time_horizon": "medium",
        "importance_score": 70,
        "confidence_score": 0.8,
        "tags": [],
        "industries": ["通信"],
        "themes": themes[item["topic_id"]],
        "stock_mentions": [],
        "key_points": [],
        "catalysts": [],
        "risks": [],
    } for item in claimed], raw_response="{}", usage={})

    deep = essay_service.deep_insights(days=30, trend_days=7)
    optical = next(item for item in deep["theme_heatmap"]["items"] if item["name"] == "光通信")

    assert optical["total"] == 3
    assert {item["name"] for item in optical["aliases"]} == {"CPO", "NPO", "光模块"}
    assert deep["theme_heatmap"]["taxonomy"]["raw_theme_count"] == 3
    assert deep["theme_heatmap"]["taxonomy"]["canonical_theme_count"] == 1
    assert deep["theme_heatmap"]["taxonomy"]["merged_theme_count"] == 2
    assert max(point["concentration_score"] for point in optical["points"]) == 100.0
    assert essay_service.list_feed(query="光通信", page=1, page_size=20)["total"] == 3


def test_large_topic_id_queries_are_split_below_sqlite_parameter_limit() -> None:
    session = MagicMock()
    session.execute.return_value.scalars.return_value.all.return_value = []

    rows = _select_in_batches(
        session,
        ResearchNote,
        ResearchNote.topic_id,
        [f"topic-{index}" for index in range(1201)],
    )

    assert rows == []
    assert session.execute.call_count == 3


def test_queue_claim_save_and_dashboard_are_restart_safe(essay_service) -> None:
    progress = essay_service.progress(days=30)
    assert progress["pending"] == 1

    claimed = essay_service.repo.claim_batch(limit=10, max_attempts=4)
    assert [item["topic_id"] for item in claimed] == ["topic-1"]
    essay_service.repo.save_successes([{
        "topic_id": "topic-1",
        "summary": "机器人订单增长，关注量产风险。",
        "primary_category": "company_research",
        "sentiment": "mixed",
        "time_horizon": "medium",
        "importance_score": 80,
        "confidence_score": 0.9,
        "tags": ["订单增长", "风险提示"],
        "industries": ["机械设备"],
        "themes": ["机器人"],
        "stock_mentions": [
            {"ts_code": "600519.SH", "name": "贵州茅台", "stance": "watching", "confidence": 0.8, "rationale": "关注量产"},
            {"ts_code": "688836.SH", "name": "宇树科技", "stance": "watching", "confidence": 0.6, "rationale": "未上市公司"},
        ],
        "key_points": ["订单增长"],
        "catalysts": ["量产"],
        "risks": ["交付延期"],
    }], raw_response="{}", usage={"prompt_tokens": 100, "completion_tokens": 50, "total_tokens": 150})

    assert essay_service.progress(days=30)["coverage_percent"] == 100.0
    dashboard = essay_service.dashboard(days=30)
    assert dashboard["summary"]["analyzed_count"] == 1
    assert dashboard["top_tags"][0] == {"name": "订单增长", "count": 1}
    assert dashboard["top_stocks"][0]["ts_code"] == "600519.SH"
    insights = essay_service.insights(days=30, trend_days=7)
    assert len(insights["trend"]) == 7
    assert insights["coverage"]["analyzed_count"] == 1
    assert insights["source_quality"][0]["name"] == "unknown"
    assert [item["name"] for item in insights["watchlist"]] == ["华懋科技", "胜宏科技"]
    deep = essay_service.deep_insights(days=30, trend_days=7)
    assert deep["summary"]["theme_count"] == 1
    assert deep["layers"]["themes"][0]["label"] == "机器人"
    deep_stocks = {item["label"]: item["ts_code"] for item in deep["layers"]["stocks"]}
    assert deep_stocks["贵州茅台"] == "600519.SH"
    assert deep_stocks["宇树科技"] == ""
    assert deep["theme_heatmap"]["items"][0]["name"] == "机器人"
    assert deep["pulse"] == []
    assert deep["period"]["start_date"] <= "2026-08-19" <= deep["period"]["end_date"]
    assert deep["evidence_funnel"][2] == {"name": "匹配有效股票代码", "count": 1}


def test_completed_placeholder_summary_is_automatically_requeued(essay_service) -> None:
    claimed = essay_service.repo.claim_batch(limit=1, max_attempts=4)
    essay_service.repo.save_successes([{
        "topic_id": claimed[0]["topic_id"],
        "summary": "信息不足，未形成有效摘要",
        "primary_category": "other",
        "sentiment": "neutral",
        "time_horizon": "unclear",
        "importance_score": 0,
        "confidence_score": 0,
        "tags": [],
        "industries": [],
        "themes": [],
        "stock_mentions": [],
        "key_points": [],
        "catalysts": [],
        "risks": [],
    }], raw_response="{}", usage={})

    repaired = essay_service.repo.requeue_low_quality_completed(
        model="deepseek-v4-flash",
        prompt_version="quality-v3",
    )

    assert repaired == 1
    with essay_service.repo.db.get_session() as session:
        row = session.execute(
            select(EssayAnalysisRecord).where(EssayAnalysisRecord.topic_id == "topic-1")
        ).scalar_one()
        assert row.status == "pending"
        assert row.attempt_count == 0
        assert row.prompt_version == "quality-v3"
        assert row.summary is None
        assert row.error_message == "正在自动重新分析"


def test_deep_insights_links_deduplicated_essay_days_to_local_prices(essay_service) -> None:
    claimed = essay_service.repo.claim_batch(limit=10, max_attempts=4)
    essay_service.repo.save_successes([{
        "topic_id": claimed[0]["topic_id"],
        "summary": "行情关联测试", "primary_category": "company_research",
        "sentiment": "bullish", "time_horizon": "short", "importance_score": 80,
        "confidence_score": 0.8, "tags": [], "industries": [], "themes": ["白酒"],
        "stock_mentions": [{"ts_code": "600519.SH", "name": "贵州茅台", "stance": "bullish", "confidence": 0.9, "rationale": "测试"}],
        "key_points": [], "catalysts": [], "risks": [],
    }], raw_response="{}", usage={})
    with essay_service.repo.db.get_session() as session:
        session.add_all([
            StockDaily(code="600519", date=date(2026, 8, 18), open=99, close=100, data_source="tushare:daily"),
            StockDaily(code="600519", date=date(2026, 8, 20), open=100, close=103, data_source="tushare:daily"),
            StockDaily(code="600519", date=date(2026, 8, 21), open=103, close=104, data_source="tushare:daily"),
            StockDaily(code="000300", date=date(2026, 8, 18), open=100, close=100, data_source="tushare:index_daily"),
            StockDaily(code="000300", date=date(2026, 8, 20), open=100, close=101, data_source="tushare:index_daily"),
            StockDaily(code="000300", date=date(2026, 8, 21), open=101, close=101, data_source="tushare:index_daily"),
        ])
        session.commit()

    deep = essay_service.deep_insights(horizon="short", end_date="2026-08-21")
    market = deep["market_impact"]
    item = market["items"][0]
    one_day = next(metric for metric in item["metrics"] if metric["period"] == 1)

    assert deep["period"] == {
        "horizon": "short", "start_date": "2026-08-08", "end_date": "2026-08-21",
        "granularity": "day", "comparison_days": 7,
    }
    assert market["dedupe_rule"].startswith("同一股票同一天")
    assert item["name"] == "贵州茅台"
    assert item["event_day_count"] == 1
    assert one_day["sample_count"] == 1
    assert one_day["average_return"] == 3.0
    assert one_day["average_excess_return"] == 2.0


def test_historical_backfill_queues_only_requested_unqueued_notes(essay_service) -> None:
    ResearchNoteService().import_topics([{
        "topic_id": "topic-old",
        "title": "较早的历史纪要",
        "content": "历史正文",
        "create_time": "2025-08-01T09:00:00+0800",
        "group": {"group_id": "g1", "name": "调研纪要"},
    }, {
        "topic_id": "topic-new",
        "title": "较新的历史纪要",
        "content": "历史正文",
        "create_time": "2025-08-02T09:00:00+0800",
        "group": {"group_id": "g1", "name": "调研纪要"},
    }], enqueue_analysis=False)

    before = essay_service.historical_backlog()
    assert before["total_notes"] == 3
    assert before["unqueued"] == 2

    queued = essay_service.enqueue_unqueued(count=1, order="newest")
    assert queued["requested"] == 1
    assert queued["selected"] == 1
    assert queued["created"] == 1
    with essay_service.repo.db.get_session() as session:
        topic_ids = set(session.execute(select(EssayAnalysisRecord.topic_id)).scalars().all())
    assert topic_ids == {"topic-1", "topic-new"}
    assert essay_service.historical_backlog()["unqueued"] == 1

    remaining = essay_service.enqueue_unqueued(count=5000, order="oldest")
    assert remaining["selected"] == 1
    assert remaining["created"] == 1
    assert essay_service.historical_backlog()["unqueued"] == 0


def test_worker_start_returns_immediately_when_already_running(monkeypatch) -> None:
    monkeypatch.setattr(
        "src.services.essay_analysis_worker.DeepSeekEssayAnalyzer",
        lambda: MagicMock(configured=True),
    )
    worker = EssayAnalysisWorker()
    worker._thread = MagicMock()
    worker._thread.is_alive.return_value = True

    result = worker.start(bootstrap_recent=False)

    assert result["running"] is True


def test_essay_radar_read_api(essay_service, monkeypatch) -> None:
    monkeypatch.setattr(
        "src.services.essay_analysis_worker.EssayAnalysisWorker.status",
        lambda self: {"running": False},
    )
    start_calls = []
    monkeypatch.setattr(
        "src.services.essay_analysis_worker.EssayAnalysisWorker.start",
        lambda self, **kwargs: start_calls.append(kwargs) or {"running": True},
    )
    app = FastAPI()
    app.include_router(essay_radar.router, prefix="/api/v1/essay-radar")
    client = TestClient(app)

    response = client.get("/api/v1/essay-radar/status")
    assert response.status_code == 200
    assert response.json()["progress"]["total_notes"] == 1

    response = client.get("/api/v1/essay-radar/analyses")
    assert response.status_code == 200
    assert response.json()["total"] == 1

    assert client.get("/api/v1/essay-radar/word-cloud?period=day&kind=stocks").status_code == 200
    assert client.get("/api/v1/essay-radar/insights?days=30&trend_days=14").status_code == 200
    assert client.get("/api/v1/essay-radar/deep-insights?days=30&trend_days=14").status_code == 200
    backlog = client.get("/api/v1/essay-radar/historical-backlog")
    assert backlog.status_code == 200
    assert backlog.json()["unqueued"] == 0
    assert backlog.json()["total_notes"] == 1
    assert backlog.json()["group_count"] == 1
    assert backlog.json()["earliest_note_at"].startswith("2026-08-19")
    assert backlog.json()["latest_note_at"].startswith("2026-08-19")
    queued = client.post("/api/v1/essay-radar/backfill-count", json={"count": 100, "order": "newest"})
    assert queued.status_code == 200
    assert queued.json()["queue"]["selected"] == 0
    assert start_calls == [{"bootstrap_recent": False}]
    assert client.post("/api/v1/essay-radar/backfill-count", json={"count": 5001, "order": "newest"}).status_code == 422
    reports = client.get("/api/v1/essay-radar/daily-reports")
    assert reports.status_code == 200
    assert "models" in reports.json()


def test_feed_searches_raw_library_without_ai_analysis(essay_service) -> None:
    ResearchNoteService().import_topics([{
        "topic_id": "topic-raw-only",
        "title": "尚未分析的历史原文",
        "content": "这段原文包含全库独家关键词和未入队信息。",
        "create_time": "2025-08-01T09:00:00+0800",
        "group": {"group_id": "g1", "name": "调研纪要"},
        "owner": {"name": "原文作者"},
    }], enqueue_analysis=False)
    app = FastAPI()
    app.include_router(essay_radar.router, prefix="/api/v1/essay-radar")
    client = TestClient(app)

    feed = client.get("/api/v1/essay-radar/feed?days=0&query=全库独家关键词")
    assert feed.status_code == 200
    assert feed.json()["scope"] == "all_stored_notes"
    assert feed.json()["total"] == 1
    assert feed.json()["items"][0]["topic_id"] == "topic-raw-only"
    assert feed.json()["items"][0]["status"] == "not_queued"
    assert "全库独家关键词" in feed.json()["items"][0]["note"]["content"]

    completed_only = client.get("/api/v1/essay-radar/feed?days=0&analysis_status=completed")
    assert completed_only.status_code == 200
    assert completed_only.json()["total"] == 0

    uncompleted = client.get("/api/v1/essay-radar/feed?days=0&analysis_status=uncompleted")
    assert uncompleted.status_code == 200
    assert uncompleted.json()["total"] == 2

    reused_total = client.get(
        "/api/v1/essay-radar/feed?days=0&query=全库独家关键词&page=2&known_total=123"
    )
    assert reused_total.status_code == 200
    assert reused_total.json()["total"] == 123


def test_feed_export_contains_analysis_columns_and_complete_long_original(essay_service) -> None:
    long_content = "超长导出原文" * 6000
    ResearchNoteService().import_topics([{
        "topic_id": "topic-export-long",
        "title": "Excel 导出完整性验收",
        "content": long_content,
        "create_time": "2026-08-20T09:00:00+0800",
        "group": {"group_id": "g1", "name": "调研纪要"},
        "owner": {"name": "导出作者"},
    }], enqueue_analysis=False)
    app = FastAPI()
    app.include_router(essay_radar.router, prefix="/api/v1/essay-radar")
    client = TestClient(app)

    response = client.get("/api/v1/essay-radar/feed/export?days=0&query=Excel%20导出完整性验收")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["x-export-row-count"] == "1"
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    assert workbook.sheetnames == ["搜索结果与分析标签", "原文全文", "导出条件"]
    result_rows = list(workbook["搜索结果与分析标签"].iter_rows(values_only=True))
    assert "标签" in result_rows[0]
    assert "AI摘要" in result_rows[0]
    assert result_rows[1][1] == "topic-export-long"
    raw_rows = list(workbook["原文全文"].iter_rows(min_row=2, values_only=True))
    assert len(raw_rows) > 1
    assert "".join(str(row[5] or "") for row in raw_rows) == long_content


def test_feed_export_selected_only_contains_checked_topics(essay_service) -> None:
    ResearchNoteService().import_topics([
        {
            "topic_id": "selected-topic",
            "title": "需要下载的小作文",
            "content": "勾选原文内容",
            "create_time": "2026-08-20T09:00:00+0800",
            "group": {"group_id": "g1", "name": "调研纪要"},
        },
        {
            "topic_id": "not-selected-topic",
            "title": "不应下载的小作文",
            "content": "未勾选内容",
            "create_time": "2026-08-20T08:00:00+0800",
            "group": {"group_id": "g1", "name": "调研纪要"},
        },
    ], enqueue_analysis=False)
    app = FastAPI()
    app.include_router(essay_radar.router, prefix="/api/v1/essay-radar")

    response = TestClient(app).post(
        "/api/v1/essay-radar/feed/export-selected",
        json={"topic_ids": ["selected-topic"]},
    )

    assert response.status_code == 200
    assert response.headers["x-export-row-count"] == "1"
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    rows = list(workbook["搜索结果与分析标签"].iter_rows(min_row=2, values_only=True))
    assert [row[1] for row in rows] == ["selected-topic"]


def test_word_cloud_and_daily_report_are_periodic_and_idempotent(essay_service, monkeypatch) -> None:
    claimed = essay_service.repo.claim_batch(limit=10, max_attempts=4)
    essay_service.repo.save_successes([{
        "topic_id": "topic-1", "summary": "机器人订单增长", "primary_category": "company_research",
        "sentiment": "bullish", "time_horizon": "medium", "importance_score": 88,
        "confidence_score": 0.9, "tags": ["订单增长"], "industries": ["机械设备"], "themes": ["机器人"],
        "stock_mentions": [{"ts_code": "688836.SH", "name": "宇树科技", "stance": "bullish", "confidence": 0.8, "rationale": "订单增长"}],
        "key_points": ["订单增长"], "catalysts": ["量产"], "risks": ["交付"],
        "evidence": [{"claim": "订单增长", "evidence": "公司称订单增长", "strength": "medium"}],
        "contradictions": [], "falsification_conditions": ["订单取消"], "monitoring_points": ["月度交付"],
        "earnings_impact": "收入可能增长", "valuation_impact": "估值取决于交付", "source_quality": "medium",
        "novelty_score": 70, "information_type": "management_guidance",
    }], raw_response="{}", usage={})

    cloud = essay_service.word_cloud(period="day", anchor_date="2026-08-19", kind="stocks")
    assert cloud["items"][0]["name"] == "宇树科技"
    assert cloud["items"][0]["count"] == 1

    def fake_call(self, model, target, rows):
        return {"executive_summary": "机器人订单是当日主线", "market_regime": "成长主题", "source_count": len(rows)}, {"total_tokens": 123}

    monkeypatch.setattr(EssayDailyReportService, "_call_model", fake_call)
    service = EssayDailyReportService(analysis_repo=essay_service.repo)
    first = service.generate(report_date="2026-08-19", models=["deepseek-v4-flash"])
    second = service.generate(report_date="2026-08-19", models=["deepseek-v4-flash"])
    assert first["models"][0]["status"] == "completed"
    assert second["models"][0]["status"] == "unchanged"
    saved = service.list(limit=5)["items"][0]
    assert saved["report"]["executive_summary"] == "机器人订单是当日主线"


def test_daily_context_uses_full_population_and_ranked_evidence() -> None:
    rows = [{
        "topic_id": f"topic-{index}",
        "summary": f"摘要 {index}",
        "primary_category": "rumor" if index == 0 else "company_research",
        "sentiment": "bullish", "importance_score": index % 100,
        "confidence_score": 0.3 if index == 0 else 0.8,
        "novelty_score": index % 90, "information_type": "market_rumor" if index == 0 else "institution_view",
        "source_quality": "low" if index == 0 else "medium", "tags": ["机器人"], "themes": ["AI"],
        "stock_mentions": [{"name": "宇树科技"}], "evidence": [{"claim": "c", "evidence": "e"}],
        "note": {"title": f"标题 {index}", "group_name": "调研纪要"},
    } for index in range(350)]
    context = EssayDailyReportService._daily_context(rows)
    assert context["coverage"]["total_records"] == 350
    assert context["coverage"]["representative_records"] == 240
    assert context["coverage"]["rumor_records"] == 1
    assert context["full_population_aggregates"]["top_stocks"][0] == ("宇树科技", 350)


def test_daily_context_keeps_low_frequency_watchlist_evidence(monkeypatch) -> None:
    monkeypatch.setenv("ESSAY_WATCHLIST", "603306.SH:华懋科技,300476.SZ:胜宏科技")
    rows = [{
        "topic_id": f"hot-{index}", "summary": "热门主题", "primary_category": "industry_chain",
        "sentiment": "bullish", "importance_score": 90, "confidence_score": 0.9,
        "novelty_score": 90, "information_type": "institution_view", "source_quality": "medium",
        "stock_mentions": [{"ts_code": "300308.SZ", "name": "中际旭创"}],
        "note": {"title": "热门算力", "group_name": "调研纪要"},
    } for index in range(300)]
    rows.append({
        "topic_id": "watch-huamao", "summary": "华懋科技跟踪", "primary_category": "company_research",
        "sentiment": "neutral", "importance_score": 10, "confidence_score": 0.6,
        "novelty_score": 5, "information_type": "fact", "source_quality": "high",
        "stock_mentions": [{"ts_code": "603306.SH", "name": "华懋科技"}],
        "note": {"title": "华懋科技更新", "group_name": "关注股星球"},
    })
    context = EssayDailyReportService._daily_context(rows)
    topic_ids = {item["topic_id"] for item in context["representative_evidence"]}
    assert "watch-huamao" in topic_ids
    assert context["coverage"]["representative_watchlist_records"] == 1
    assert context["coverage"]["selection_strategy"].startswith("watchlist_priority")


def test_cross_model_comparison_uses_structured_theme_and_stock_directions() -> None:
    rows = [
        {"model": "model-a", "report": {
            "key_themes": [{"name": "AI算力", "direction": "bullish"}],
            "stock_focus": [{"ts_code": "300476.SZ", "name": "胜宏科技", "stance": "bullish"}],
            "consensus": ["算力景气延续"], "divergences": [],
        }},
        {"model": "model-b", "report": {
            "key_themes": [{"name": "AI算力", "direction": "bullish"}],
            "stock_focus": [{"ts_code": "300476.SZ", "name": "胜宏科技", "stance": "bearish"}],
            "consensus": ["算力的景气仍在延续"], "divergences": [],
        }},
    ]
    comparison = EssayAnalysisService._compare_reports(rows)
    assert any("AI算力" in item["text"] and item["model_count"] == 2 for item in comparison["consensus"])
    assert any("胜宏科技" in item["text"] and item["model_count"] == 2 for item in comparison["divergences"])


def test_dashboard_and_insights_share_decoded_row_snapshot(essay_service, monkeypatch) -> None:
    calls = 0
    rows = [{"topic_id": "cached-row"}]

    def fake_rows(_repo, *, cutoff):
        nonlocal calls
        calls += 1
        return rows

    _dashboard_rows_cache.clear()
    monkeypatch.setattr(EssayAnalysisRepository, "completed_for_dashboard", fake_rows)
    assert essay_service._completed_dashboard_rows(30) is rows
    assert essay_service._completed_dashboard_rows(30) is rows
    assert calls == 1
    _dashboard_rows_cache.clear()
