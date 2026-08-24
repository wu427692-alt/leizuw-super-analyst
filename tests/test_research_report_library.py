from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from src.services.research_report_library_service import ResearchReportLibraryService
from src.storage import DatabaseManager


class FakeFinancial:
    def __init__(self):
        self.calls = []

    def query(self, *, source, resource, params, fields=None):
        self.calls.append({"source": source, "resource": resource, "params": params, "fields": fields})
        if params["offset"]:
            return {"rows": []}
        return {"rows": [
            {
                "trade_date": "20260820",
                "title": "低空经济产业链深度研究",
                "abstr": "无人机、eVTOL与低空基础设施进入商业化阶段。",
                "report_type": "行业研报",
                "author": "研究员甲",
                "name": "",
                "ts_code": "",
                "inst_csname": "测试证券",
                "ind_name": "航空装备",
                "url": "https://example.com/low-altitude.pdf",
            },
            {
                "trade_date": "20260819",
                "title": "华懋科技公司点评",
                "abstr": "安全气囊材料业务跟踪。",
                "report_type": "个股研报",
                "author": "研究员乙",
                "name": "华懋科技",
                "ts_code": "603306.SH",
                "inst_csname": "示例证券",
                "ind_name": "汽车零部件",
                "url": "https://example.com/603306.pdf",
            },
        ]}


def test_two_year_report_library_sync_search_facets_and_export(tmp_path):
    DatabaseManager.reset_instance()
    db = DatabaseManager(db_url=f"sqlite:///{tmp_path / 'reports.db'}")
    financial = FakeFinancial()
    service = ResearchReportLibraryService(db_manager=db, financial=financial)
    try:
        windows = service._date_windows(date(2026, 8, 18), date(2026, 8, 24))
        service._write_state(
            status="queued", progress=0, start_date=date(2026, 8, 18), end_date=date(2026, 8, 24),
            total_windows=len(windows), completed_windows=0, scanned_rows=0, saved_rows=0,
        )
        service._run_sync(date(2026, 8, 18), date(2026, 8, 24), windows)

        status = service.status()
        assert status["status"] == "completed"
        assert status["total"] == 2
        assert status["pdf_count"] == 2
        assert financial.calls[0]["resource"] == "research_report"
        assert "abstr" in financial.calls[0]["fields"]

        by_title = service.search(title_query="低空经济", broker="测试证券")
        assert by_title["total"] == 1
        assert by_title["items"][0]["title"] == "低空经济产业链深度研究"
        assert "深度研究" in by_title["items"][0]["tags"]

        by_content = service.search(content_query="安全气囊", company="华懋科技")
        assert by_content["total"] == 1
        assert by_content["items"][0]["ts_code"] == "603306.SH"

        facets = service.facets()
        assert {item["value"] for item in facets["brokers"]} == {"测试证券", "示例证券"}
        assert any(item["value"] == "深度研究" for item in facets["tags"])

        export = service.export_selected([by_title["items"][0]["id"]])
        workbook = load_workbook(BytesIO(export), read_only=True)
        sheet = workbook["已选研报"]
        assert sheet.max_row == 2
        assert sheet.cell(2, 2).value == "低空经济产业链深度研究"
        assert sheet.cell(2, 11).value == "https://example.com/low-altitude.pdf"
    finally:
        service._executor.shutdown(wait=True, cancel_futures=True)
        DatabaseManager.reset_instance()

